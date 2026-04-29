"""Shared helper for extracting Customer / Project / PO / SO metadata from
previously-generated ATLAS workbooks.

Used by both the Packing Slip frame (when a CSV/XLSX file is uploaded for slip
generation) and the Inventory frame (when an existing inventory report is
selected for append mode), so the auto-populate behaviour stays consistent
across both modes.

Tries two known formats in order:

  1. Packing-slip / per-device sheet — header cells live at C5/C6/C7/D7.
  2. Inventory-report summary sheet — values live at B7/D7.

Returns blank strings for any field that can't be located. Never raises.
"""
import logging
from typing import Dict


def extract_workbook_metadata(file_path: str) -> Dict[str, str]:
    """Extract Customer/Project/PO/SO from an existing ATLAS workbook.

    Args:
        file_path: Absolute path to a .xlsx/.xls file.

    Returns:
        Dict with keys "customer", "project", "po", "so". Any field that
        cannot be parsed is returned as an empty string. Returns all-empty
        dict if the file cannot be opened.
    """
    result = {"customer": "", "project": "", "po": "", "so": ""}
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    except Exception as e:
        logging.debug(f"[METADATA] Could not open workbook for metadata: {e}")
        return result

    def _clean(val) -> str:
        if val is None:
            return ""
        s = str(val).strip()
        if s.lower() in ("", "nan", "none"):
            return ""
        return s

    try:
        device_sheets = [n for n in wb.sheetnames if "summary" not in n.lower()]
        summary_sheets = [n for n in wb.sheetnames if "summary" in n.lower()]

        # Strategy 1: per-device sheet headers (C5/C6/C7/D7).
        if device_sheets:
            ws = wb[device_sheets[0]]
            for coord, key in (("C5", "customer"), ("C6", "project"),
                                ("C7", "po"), ("D7", "so")):
                if not result[key]:
                    result[key] = _clean(ws[coord].value)

        # Strategy 2: summary sheet (B7 = customer, D7 = project).
        if summary_sheets:
            ws = wb[summary_sheets[0]]
            if not result["customer"]:
                result["customer"] = _clean(ws["B7"].value)
            if not result["project"]:
                result["project"] = _clean(ws["D7"].value)
            # Some inventory templates also store PO/SO on the summary sheet.
            if not result["po"]:
                result["po"] = _clean(ws["F7"].value)
            if not result["so"]:
                result["so"] = _clean(ws["H7"].value)
    except Exception as e:
        logging.debug(f"[METADATA] Error reading workbook metadata: {e}")
    finally:
        try:
            wb.close()
        except Exception:
            pass

    return result
