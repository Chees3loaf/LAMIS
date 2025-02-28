import os
import sys
import pandas as pd
import openpyxl
from tkinter import simpledialog, messagebox
import logging

logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')

# Add the parent directory to the Python path
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

class PackingSlipGenerator:
    def __init__(self, template_path, save_location):
        self.template_path = template_path
        self.save_location = save_location

    def create_packing_slips(self, file_path, ip_list):
        logging.debug(f"Loading Excel file from {file_path}")
        df = pd.read_excel(file_path)

        # Prompt user for packing slip information
        customer = simpledialog.askstring("Input", "Enter the customer:")
        project = simpledialog.askstring("Input", "Enter the project:")
        
        sale_order_numbers = []
        while True:
            sale_order_number = simpledialog.askstring("Input", "Enter a sales order number (or leave blank to finish):")
            if not sale_order_number:
                break
            sale_order_numbers.append(sale_order_number)
        
        customer_po_numbers = []
        while True:
            customer_po_number = simpledialog.askstring("Input", "Enter a customer PO number (or leave blank to finish):")
            if not customer_po_number:
                break
            customer_po_numbers.append(customer_po_number)

        logging.debug("Generating packing slips for each device.")
        for i, ip in enumerate(ip_list):
            try:
                logging.debug(f"Loading template for IP: {ip}")
                wb = openpyxl.load_workbook(self.template_path)
                ws = wb.active

                # Fill out the packing slip
                ws['B1'] = customer
                ws['B2'] = project
                ws['B3'] = sale_order_numbers[i % len(sale_order_numbers)]
                ws['B4'] = customer_po_numbers[i % len(customer_po_numbers)]

                # Fill the data from the inventory file
                device_data = df[df['IP Address'] == ip]
                for idx, row in device_data.iterrows():
                    ws.append([row['Description'], row['Part Number'], row['Serial Number']])

                # Save the packing slip
                save_path = os.path.join(self.save_location, f"Packing Slip - {ip}.xlsx")
                wb.save(save_path)
                logging.debug(f"Packing slip saved to {save_path}")
            except Exception as e:
                logging.error(f"Failed to create packing slip for IP: {ip} - {e}")

        messagebox.showinfo("Packing Slips", "Packing slips have been created successfully.")

if __name__ == "__main__":
    # Example usage
    template_path = '/mnt/data/LAMIS_Packing_slip.xlsx'
    save_location = '/mnt/data/generated_slips'
    generator = PackingSlipGenerator(template_path, save_location)
    generator.create_packing_slips('/mnt/data/Device_report_template.xlsx', ['192.168.0.1', '192.168.0.2'])
