from pathlib import Path
from openpyxl import load_workbook

wb_path = Path(r"C:\Users\ZackerySimino\OneDrive - LightRiver Technologies Inc\Desktop\Inventory and Packing slips\BPA\TO3\Master_BPA_Task_Order_3_Raw_Report_2026-05-10_20-49.xlsx")

if not wb_path.exists():
    raise SystemExit(f"Workbook not found: {wb_path}")

wb = load_workbook(wb_path, data_only=True)

print(f"Workbook: {wb_path.name}")
print(f"Total sheets: {len(wb.sheetnames)}")
print(f"Device sheets (excluding Summary): {len([s for s in wb.sheetnames if s != 'Summary'])}")
print()

sheet_name = "CUSI002_7705"
if sheet_name not in wb.sheetnames:
    raise SystemExit(f"Missing sheet: {sheet_name}")

ws = wb[sheet_name]

target_rows = []
for r_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
    for value in row:
        text = "" if value is None else str(value)
        if "3HE06792AA" in text:
            target_rows.append((r_idx, row))
            break

if not target_rows:
    print("No row found containing 3HE06792AA in CUSI002_7705")
else:
    print("Rows in CUSI002_7705 containing 3HE06792AA:")
    for r_idx, row in target_rows:
        cols = [f"C{idx + 1}={repr(val)}" for idx, val in enumerate(row)]
        print(f"  Row {r_idx}: " + ", ".join(cols))

print()
expected_desc = "Fan Module (SAR-8 Shelf V2) Ext Temp"
found_expected = False
for sname in wb.sheetnames:
    if sname == "Summary":
        continue
    ws2 = wb[sname]
    for row in ws2.iter_rows(values_only=True):
        if any((str(v) == expected_desc) for v in row if v is not None):
            found_expected = True
            break
    if found_expected:
        break

print(f"Expected description present anywhere: {found_expected}")

# Count literal 'Not Found' cells across workbook for quick trend check
not_found_cells = 0
for sname in wb.sheetnames:
    if sname == "Summary":
        continue
    ws2 = wb[sname]
    for row in ws2.iter_rows(values_only=True):
        for value in row:
            if value is not None and str(value).strip() == "Not Found":
                not_found_cells += 1

print(f"Literal 'Not Found' cell count (all columns): {not_found_cells}")
