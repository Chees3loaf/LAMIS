"""
gui/raw_frame.py - "Raw" mode: process manually captured CLI output files.

Supported input formats
-----------------------
* **.txt** — single device, raw CLI transcript in plain text.
* **.xlsx / .xls** — one *or* many devices:
  - Single sheet  → treated as one device (sheet name used as Device ID).
  - Multi-sheet   → each sheet is a separate device; all results are merged
                    into a single Device Report workbook.

In each case ATLAS splits the transcript by command boundaries, feeds the
sections into the normal process_outputs() pipeline, and exports a Device
Report using the same templates as a live network scan.
"""
import importlib
import logging
import os
import re
import threading
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
import tkinter as tk
from tkinter import ttk, scrolledtext, filedialog, messagebox

# Display name → module path for the script dropdown
SCRIPT_OPTIONS: Dict[str, str] = {
    "Nokia PSI":   "scripts.Nokia_PSI",
    "Nokia 1830":  "scripts.Nokia_1830",
    "Nokia SAR":   "scripts.Nokia_SAR",
    "Nokia IXR":   "scripts.Nokia_IXR",
    "Ciena 6500":  "scripts.Ciena_6500",
    "Ciena RLS":   "scripts.Ciena_RLS",
}

# Map module path → workbook-builder family key
_FAMILY_BY_MODULE: Dict[str, str] = {
    "scripts.Nokia_PSI": "psi",
    "scripts.Ciena_RLS": "rls",
}


def _split_raw_output_by_commands(raw_text: str, commands: List[str]) -> List[str]:
    """Split a raw CLI transcript into per-command output sections.

    The transcript may optionally have hostname prompts on command lines,
    e.g.::

        USDEN5-L9O1# show shelf 1
        ...output...
        USDEN5-L9O1# show card inventory *
        ...output...

    or bare lines like ``show shelf 1``.

    Returns a list with one entry per command (parallel to *commands*).
    An empty string is returned for any command not found in the transcript.
    Output lines that appear *before* the first recognised command (e.g. a
    login banner) are discarded.
    """
    lines = raw_text.splitlines()

    # Find the first line index where each command appears.
    cmd_line: Dict[int, int] = {}  # command_index -> line_index

    for line_idx, line in enumerate(lines):
        # Strip a leading "HOSTNAME# " prompt if present, then compare.
        candidate = re.sub(r"^[^\#]*#\s*", "", line).strip()
        for cmd_idx, cmd in enumerate(commands):
            if cmd_idx not in cmd_line and candidate.startswith(cmd):
                cmd_line[cmd_idx] = line_idx
                break  # a line can only start one command

    sorted_positions = sorted(cmd_line.items(), key=lambda x: x[1])
    result = [""] * len(commands)

    for i, (cmd_idx, start_line) in enumerate(sorted_positions):
        # Output is everything *after* the command line up to the next command.
        end_line = (
            sorted_positions[i + 1][1] if i + 1 < len(sorted_positions) else len(lines)
        )
        result[cmd_idx] = "\n".join(lines[start_line + 1 : end_line])

    return result


def _read_excel_sheets(file_path: str) -> Dict[str, str]:
    """Read an Excel workbook and return ``{sheet_name: raw_text}`` for every sheet.

    Each sheet is expected to have the CLI transcript in column A, one line
    per row (this is the format produced by Nokia's CPAM / manual capture
    tools).  Blank rows are preserved as empty lines so command-boundary
    detection still works.
    """
    try:
        import openpyxl
    except ImportError as exc:
        raise RuntimeError("openpyxl is required to read Excel files.") from exc

    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    sheets: Dict[str, str] = {}
    for name in wb.sheetnames:
        ws = wb[name]
        lines = []
        for row in ws.iter_rows(values_only=True):
            cell_val = row[0] if row else None
            lines.append(str(cell_val) if cell_val is not None else "")
        sheets[name] = "\n".join(lines)
    wb.close()
    return sheets


class RawFrame(ttk.Frame):
    """UI panel for the 'Raw' mode."""

    def __init__(self, parent: tk.Widget, gui: Any) -> None:
        super().__init__(parent)
        self.gui = gui  # reference to InventoryGUI instance
        self._file_path: Optional[str] = None
        self._running = False
        self._setup_ui()

    # ------------------------------------------------------------------
    # Widget construction
    # ------------------------------------------------------------------

    def _setup_ui(self) -> None:
        pad: Dict[str, int] = {"padx": 8, "pady": 4}

        # ── File upload row ──────────────────────────────────────────
        file_frame = ttk.LabelFrame(
            self,
            text="Input File  (.txt plain text  OR  .xlsx/.xls multi-tab workbook — one device per sheet)",
        )
        file_frame.pack(fill=tk.X, **pad)

        self._file_label = tk.StringVar(value="No file selected")
        ttk.Label(
            file_frame, textvariable=self._file_label, width=65, anchor="w"
        ).pack(side=tk.LEFT, padx=6, pady=6)
        ttk.Button(
            file_frame, text="Browse…", command=self._browse_file
        ).pack(side=tk.LEFT, padx=6, pady=6)

        # ── Script + device label row ────────────────────────────────
        cfg_frame = ttk.LabelFrame(self, text="Processing Options")
        cfg_frame.pack(fill=tk.X, **pad)

        ttk.Label(cfg_frame, text="Device Type:").grid(
            row=0, column=0, padx=(8, 4), pady=6, sticky="w"
        )
        self._script_var = tk.StringVar(value="Nokia PSI")
        script_cb = ttk.Combobox(
            cfg_frame,
            textvariable=self._script_var,
            values=list(SCRIPT_OPTIONS.keys()),
            state="readonly",
            width=18,
        )
        script_cb.grid(row=0, column=1, padx=4, pady=6, sticky="w")

        ttk.Label(
            cfg_frame,
            text="Device ID  (single .txt only — ignored for multi-sheet Excel):",
        ).grid(row=0, column=2, padx=(24, 4), pady=6, sticky="w")
        self._device_id_var = tk.StringVar()
        ttk.Entry(cfg_frame, textvariable=self._device_id_var, width=24).grid(
            row=0, column=3, padx=4, pady=6, sticky="w"
        )

        # ── Sheet summary label (populated after Excel load) ─────────
        self._sheet_info_var = tk.StringVar(value="")
        ttk.Label(cfg_frame, textvariable=self._sheet_info_var, foreground="steelblue").grid(
            row=1, column=0, columnspan=4, padx=8, pady=(0, 4), sticky="w"
        )

        # ── Control bar ──────────────────────────────────────────────
        ctrl_frame = ttk.Frame(self)
        ctrl_frame.pack(fill=tk.X, **pad)

        self._run_btn = ttk.Button(
            ctrl_frame, text="Process File", command=self._on_run
        )
        self._run_btn.pack(side=tk.LEFT, padx=6)

        self._status_var = tk.StringVar(value="Ready — select a file and click Process File")
        ttk.Label(ctrl_frame, textvariable=self._status_var, foreground="gray").pack(
            side=tk.LEFT, padx=12
        )

        # ── Log window ───────────────────────────────────────────────
        log_frame = ttk.LabelFrame(self, text="Processing Log")
        log_frame.pack(fill=tk.BOTH, expand=True, **pad)
        self._log = scrolledtext.ScrolledText(
            log_frame, height=14, state="disabled", wrap="word"
        )
        self._log.pack(fill=tk.BOTH, expand=True, padx=4, pady=4)

    # ------------------------------------------------------------------
    # Callbacks
    # ------------------------------------------------------------------

    def _browse_file(self) -> None:
        path = filedialog.askopenfilename(
            title="Select device CLI output file",
            filetypes=[
                ("Supported files", "*.txt *.xlsx *.xls"),
                ("Text files", "*.txt"),
                ("Excel workbooks", "*.xlsx *.xls"),
                ("All files", "*.*"),
            ],
        )
        if not path:
            return
        self._file_path = path
        ext = Path(path).suffix.lower()
        self._log_clear()
        self._sheet_info_var.set("")

        if ext in (".xlsx", ".xls"):
            try:
                import openpyxl
                wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
                names = wb.sheetnames
                wb.close()
                count = len(names)
                preview = ", ".join(names[:6]) + ("…" if count > 6 else "")
                self._file_label.set(f"{Path(path).name}  [{count} sheet(s)]")
                self._sheet_info_var.set(
                    f"Sheets ({count}): {preview}  —  each sheet will be processed as a separate device"
                    if count > 1
                    else f"Single sheet: {names[0]}"
                )
                self._status_var.set(
                    f"Excel file loaded — {count} device(s) detected. Click Process File."
                )
            except Exception as exc:
                self._file_label.set(f"{Path(path).name}  [⚠ could not read sheets]")
                self._status_var.set(f"Warning: {exc}")
        else:
            self._file_label.set(Path(path).name)
            self._status_var.set("Text file loaded — click Process File to begin")

    def _on_run(self) -> None:
        if self._running:
            return
        if not self._file_path or not os.path.isfile(self._file_path):
            messagebox.showerror("No File", "Please select an input file first.")
            return
        script_name = self._script_var.get()
        if script_name not in SCRIPT_OPTIONS:
            messagebox.showerror("Invalid Script", f"Unknown script: {script_name!r}")
            return

        device_id = self._device_id_var.get().strip() or Path(self._file_path).stem

        self._run_btn.configure(state="disabled")
        self._running = True
        self._status_var.set("Processing…")
        self._log_clear()

        threading.Thread(
            target=self._worker,
            args=(self._file_path, device_id, script_name),
            daemon=True,
        ).start()

    # ------------------------------------------------------------------
    # Background worker
    # ------------------------------------------------------------------

    def _worker(self, file_path: str, device_id: str, script_name: str) -> None:
        """Load the file, detect format, and dispatch to single- or multi-device processing."""
        try:
            ext = Path(file_path).suffix.lower()

            if ext in (".xlsx", ".xls"):
                sheets = _read_excel_sheets(file_path)
                self._log_write(
                    f"Excel workbook loaded: {Path(file_path).name}\n"
                    f"Sheets found ({len(sheets)}): {', '.join(sheets.keys())}\n\n"
                )
                if len(sheets) == 1:
                    name, text = next(iter(sheets.items()))
                    self._process_single(text, name, script_name)
                else:
                    self._process_multi(sheets, script_name)
            else:
                self._log_write(f"Reading file: {file_path}\n")
                raw_text = Path(file_path).read_text(encoding="utf-8", errors="replace")
                self._log_write(
                    f"File loaded — {len(raw_text):,} characters, "
                    f"{raw_text.count(chr(10)):,} lines\n\n"
                )
                self._process_single(raw_text, device_id, script_name)

        except Exception as exc:
            logging.exception("Raw processing worker error")
            self._log_write(f"\n⚠  ERROR: {exc}\n")
            self.after(0, self._finish, False)

    def _process_single(self, raw_text: str, device_id: str, script_name: str) -> None:
        """Parse one device's transcript and hand off to export."""
        outputs: Dict[str, Any] = {}
        ok = self._parse_device(raw_text, device_id, script_name, outputs)
        if not ok:
            self.after(0, self._finish, False)
            return
        # Re-key the single device to "Manual" so the summary IP column is clean.
        rekeyed = {"Manual": outputs.pop(device_id)} if device_id in outputs else outputs
        family = _FAMILY_BY_MODULE.get(SCRIPT_OPTIONS[script_name], "default")
        self.after(0, self._export, rekeyed, family, device_id)

    def _process_multi(self, sheets: Dict[str, str], script_name: str) -> None:
        """Parse every sheet as its own device, merge results, then export."""
        raw_outputs: Dict[str, Any] = {}
        success_count = 0

        for sheet_name, raw_text in sheets.items():
            self._log_write(f"{'─'*50}\n")
            self._log_write(f"Processing sheet: {sheet_name!r}\n")
            ok = self._parse_device(raw_text, sheet_name, script_name, raw_outputs)
            if ok:
                success_count += 1

        self._log_write(f"\n{'═'*50}\n")
        self._log_write(
            f"Multi-sheet complete: {success_count}/{len(sheets)} device(s) parsed successfully.\n"
        )

        if success_count == 0:
            self._log_write("⚠  No devices produced data. Aborting export.\n")
            self.after(0, self._finish, False)
            return

        # Re-key every device as "Manual" / "Manual_02" / "Manual_03" … so the
        # summary IP column shows "Manual" instead of the node hostname.
        # Zero-pad numbers so lexicographic sort matches numeric order.
        merged_outputs: Dict[str, Any] = {}
        n = len(raw_outputs)
        width = len(str(n)) if n > 1 else 1
        for i, (device_id, data) in enumerate(raw_outputs.items(), start=1):
            key = "Manual" if i == 1 else f"Manual_{str(i).zfill(width)}"
            merged_outputs[key] = data

        family = _FAMILY_BY_MODULE.get(SCRIPT_OPTIONS[script_name], "default")
        label = Path(self._file_path).stem if self._file_path else "MultiDevice"
        self.after(0, self._export, merged_outputs, family, label)

    def _parse_device(
        self,
        raw_text: str,
        device_id: str,
        script_name: str,
        outputs: Dict[str, Any],
    ) -> bool:
        """Parse one device transcript into *outputs*.  Returns True on success."""
        try:
            module_path = SCRIPT_OPTIONS[script_name]
            mod = importlib.import_module(module_path)
            script_inst = mod.Script(
                ip_address=device_id,
                connection_type="ssh",
                db_cache=self.gui.db_cache,
                db_path=self.gui.db_file,
            )

            commands = script_inst.get_commands()
            outputs_list = _split_raw_output_by_commands(raw_text, commands)

            found = sum(1 for o in outputs_list if o.strip())
            self._log_write(
                f"  {found}/{len(commands)} command sections matched"
                f" for {device_id!r}\n"
            )
            for cmd, out in zip(commands, outputs_list):
                n = len(out.strip().splitlines()) if out.strip() else 0
                status = f"{n} lines" if n else "not found"
                self._log_write(f"    {cmd!r}: {status}\n")

            if found == 0:
                self._log_write(f"  ⚠  Skipping {device_id!r} — no sections matched.\n")
                return False

            script_inst.process_outputs(outputs_list, device_id, outputs)

            # Patch every DataFrame for this device so that:
            #   • System Name → device_id  (drives Device Name column + tab title)
            #   • Source       → "Manual"   (drives the IP/Source field on device sheets)
            if device_id in outputs and isinstance(outputs[device_id], dict):
                for section_data in outputs[device_id].values():
                    if not isinstance(section_data, dict):
                        continue
                    df = section_data.get("DataFrame")
                    if df is not None and hasattr(df, "__setitem__"):
                        try:
                            df["System Name"] = device_id
                            df["Source"] = "Manual"
                        except Exception:
                            pass
                    # Also patch System Info dict so workbook_builder fallbacks agree
                    si = section_data.get("System Info")
                    if isinstance(si, dict):
                        si["System Name"] = device_id
                        si["Source"] = "Manual"

            return True

        except Exception as exc:
            logging.exception(f"Error parsing device {device_id!r}")
            self._log_write(f"  ⚠  Error parsing {device_id!r}: {exc}\n")
            return False

    # ------------------------------------------------------------------
    # Export  (always called on the main thread via self.after())
    # ------------------------------------------------------------------

    def _export(self, outputs: Dict[str, Any], family: str, device_id: str) -> None:
        try:
            self._log_write("\nCollecting project information…\n")
            default_name = re.sub(r"[^\w\-]", "_", device_id)
            user_in = self.gui.get_user_inputs(default_name)
            if not user_in:
                self._log_write("Export cancelled.\n")
                self._finish(False)
                return

            customer = user_in.get("Customer", "")
            project  = user_in.get("Project", "")
            po       = user_in.get("Purchase Order", "")
            so       = user_in.get("Sales Order", "")
            filename = user_in.get("Filename", default_name)

            from utils.helpers import sanitize_filename_component
            timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M")
            safe_name = "_".join(
                filter(None, [
                    sanitize_filename_component(filename),
                    sanitize_filename_component(customer),
                    sanitize_filename_component(project),
                    "Raw_Report",
                    timestamp,
                ])
            )

            save_dir = filedialog.askdirectory(title="Choose folder to save report")
            if not save_dir:
                self._log_write("Export cancelled (no folder chosen).\n")
                self._finish(False)
                return

            output_file = os.path.join(save_dir, f"{safe_name}.xlsx")
            self._log_write(f"Saving report to:\n  {output_file}\n")
            self._status_var.set("Exporting…")

            if family == "psi":
                self.gui.build_psi_report_workbook(
                    outputs, output_file,
                    customer=customer, project=project,
                    customer_po=po, sales_order=so,
                )
            elif family == "rls":
                self.gui.build_unified_report_workbook(
                    {"rls": outputs}, output_file,
                    customer=customer, project=project,
                    customer_po=po, sales_order=so,
                )
            else:
                self.gui.build_report_workbook(
                    outputs, output_file,
                    customer=customer, project=project,
                    customer_po=po, sales_order=so,
                )

            self._log_write("✓ Export complete!\n")
            self._finish(True)

        except Exception as exc:
            logging.exception("Raw export error")
            self._log_write(f"\n⚠  Export ERROR: {exc}\n")
            self._finish(False)

    def _finish(self, success: bool) -> None:
        self._running = False
        self._run_btn.configure(state="normal")
        self._status_var.set(
            "✓ Done — report saved!" if success else "⚠ Error — see log above"
        )

    # ------------------------------------------------------------------
    # Log helpers  (thread-safe via self.after())
    # ------------------------------------------------------------------

    def _log_write(self, text: str) -> None:
        def _write() -> None:
            self._log.configure(state="normal")
            self._log.insert(tk.END, text)
            self._log.see(tk.END)
            self._log.configure(state="disabled")
        try:
            self.after(0, _write)
        except Exception:
            pass  # widget may already be destroyed

    def _log_clear(self) -> None:
        self._log.configure(state="normal")
        self._log.delete("1.0", tk.END)
        self._log.configure(state="disabled")
