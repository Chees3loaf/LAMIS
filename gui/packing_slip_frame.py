"""Packing slip frame — widgets and logic for the Packing Slip (From File) mode."""
import os
import shutil
import tempfile
import logging
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from typing import Dict, List

import openpyxl
import pandas as pd


class PackingSlipFrame(ttk.Frame):
    """Tkinter frame for the Packing Slip (From File) mode.

    Owns file-upload widgets, project information fields, and all packing slip
    generation logic for file-based input.  Uses *controller* to access
    ``controller.output_screen`` and ``controller.workbook_builder``.
    """

    def __init__(self, parent: ttk.Frame, controller) -> None:
        super().__init__(parent)
        self.controller = controller
        self.uploaded_file_path: str | None = None
        self.uploaded_file_data: pd.DataFrame | None = None
        self._multisheet_device_file: bool = False
        self._last_customer: str = ""
        self._last_project: str = ""
        self._last_customer_po: str = ""
        self._last_sales_order: str = ""
        self._build()

    # ------------------------------------------------------------------
    # Frame construction
    # ------------------------------------------------------------------

    def _build(self) -> None:
        file_frame = ttk.LabelFrame(self, text="Upload File")
        file_frame.pack(fill=tk.X, pady=5)

        self.file_path_label = tk.Label(file_frame, text="No file selected", foreground="gray")
        self.file_path_label.pack(side=tk.LEFT, padx=10, pady=5)

        tk.Button(file_frame, text="Browse", command=self.upload_file).pack(side=tk.LEFT, padx=5)

        info_frame = ttk.LabelFrame(self, text="Project Information")
        info_frame.pack(fill=tk.X, padx=5, pady=5)

        tk.Label(info_frame, text="Customer:").grid(row=0, column=0, sticky="w", padx=5, pady=5)
        self.ps_customer_entry = tk.Entry(info_frame, width=40)
        self.ps_customer_entry.grid(row=0, column=1, padx=5, pady=5)

        tk.Label(info_frame, text="Project:").grid(row=1, column=0, sticky="w", padx=5, pady=5)
        self.ps_project_entry = tk.Entry(info_frame, width=40)
        self.ps_project_entry.grid(row=1, column=1, padx=5, pady=5)

        tk.Label(info_frame, text="Purchase Order:").grid(row=2, column=0, sticky="w", padx=5, pady=5)
        self.ps_po_entry = tk.Entry(info_frame, width=40)
        self.ps_po_entry.grid(row=2, column=1, padx=5, pady=5)

        tk.Label(info_frame, text="Sales Order:").grid(row=3, column=0, sticky="w", padx=5, pady=5)
        self.ps_so_entry = tk.Entry(info_frame, width=40)
        self.ps_so_entry.grid(row=3, column=1, padx=5, pady=5)

        ps_control_frame = ttk.Frame(self)
        ps_control_frame.pack(fill=tk.X, pady=10)

        self.ps_run_button = tk.Button(
            ps_control_frame, text="Generate Packing Slips",
            command=self.generate_packing_slips_from_file,
        )
        self.ps_run_button.pack(side=tk.RIGHT, padx=5)

        self.ps_status_label = tk.Label(ps_control_frame, text="Status: Ready", anchor="w")
        self.ps_status_label.pack(side=tk.RIGHT, padx=10)

    # ------------------------------------------------------------------
    # File upload
    # ------------------------------------------------------------------

    def upload_file(self) -> None:
        """Prompt for a CSV or Excel file and load it into memory."""
        file_path = filedialog.askopenfilename(
            title="Select File (CSV or Excel)",
            filetypes=[
                ("Excel files", "*.xlsx *.xls"),
                ("CSV files", "*.csv"),
                ("All files", "*.*"),
            ],
        )
        if not file_path:
            return

        try:
            self.uploaded_file_path = file_path
            if file_path.endswith(".csv"):
                self.uploaded_file_data = pd.read_csv(file_path)
                self._multisheet_device_file = False
                count_label = f"{len(self.uploaded_file_data)} rows"
            elif file_path.endswith((".xlsx", ".xls")):
                xl = pd.ExcelFile(file_path)
                sheet_names = xl.sheet_names
                has_summary = any("summary" in s.lower() for s in sheet_names)
                device_sheets = [s for s in sheet_names if "summary" not in s.lower()]
                if len(sheet_names) > 1 and device_sheets:
                    # Multi-sheet file: each non-summary sheet is one device.
                    # This handles both plain device-report files (no summary)
                    # AND inventory reports (has Summary + device sheets).
                    # _process_multisheet_device_file will auto-skip any summary
                    # sheet because it has no PART NUMBER / SERIAL NUMBER header.
                    self._multisheet_device_file = True
                    self.uploaded_file_data = pd.DataFrame()  # placeholder
                    count_label = f"{len(device_sheets)} device(s)"
                else:
                    self._multisheet_device_file = False
                    self.uploaded_file_data = pd.read_excel(file_path)
                    count_label = f"{len(self.uploaded_file_data)} rows"
            else:
                messagebox.showerror("File Error", "Unsupported file format. Please use CSV or Excel files.")
                return

            if file_path.endswith((".xlsx", ".xls")):
                self._try_populate_fields_from_file(file_path)

            file_name = os.path.basename(file_path)
            self.file_path_label.config(
                text=f"✓ {file_name} ({count_label})",
                foreground="green",
            )
            out = self.controller.output_screen
            out.insert(tk.END, f"Loaded file: {file_name} with {count_label}\n")
            out.see(tk.END)
            logging.info(f"File uploaded: {file_path} with {len(self.uploaded_file_data)} rows")

        except Exception as e:
            messagebox.showerror("File Error", f"Failed to load file:\n{e}")
            logging.error(f"File upload error: {e}")

    # ------------------------------------------------------------------
    # Metadata auto-populate
    # ------------------------------------------------------------------

    def _try_populate_fields_from_file(self, file_path: str) -> None:
        """Read Customer, Project, PO, and SO from a previously-generated
        packing slip (or inventory report) workbook and pre-fill the form
        fields so the user can review and edit before generating."""
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            customer, project, po, so = "", "", "", ""

            device_sheets = [n for n in wb.sheetnames if "summary" not in n.lower()]
            summary_sheets = [n for n in wb.sheetnames if "summary" in n.lower()]

            # Strategy 1: packing slip / device report format
            # Header cells: C5 = Customer, C6 = Project, C7 = PO, C8 = SO
            if device_sheets:
                ws = wb[device_sheets[0]]
                for coord, target in (
                    ("C5", "customer"), ("C6", "project"),
                    ("C7", "po"), ("D7", "so"),
                ):
                    val = ws[coord].value
                    if val and str(val).strip() not in ("", "nan", "None"):
                        if target == "customer":
                            customer = str(val).strip()
                        elif target == "project":
                            project = str(val).strip()
                        elif target == "po":
                            po = str(val).strip()
                        elif target == "so":
                            so = str(val).strip()

            # Strategy 2: inventory report summary sheet
            # Summary sheet: B7 = Customer value, D7 = Project value
            if not customer and summary_sheets:
                ws = wb[summary_sheets[0]]
                b7 = ws["B7"].value
                d7 = ws["D7"].value
                if b7 and str(b7).strip() not in ("", "nan", "None"):
                    customer = str(b7).strip()
                if d7 and str(d7).strip() not in ("", "nan", "None"):
                    project = str(d7).strip()

            wb.close()

            # Pre-fill each field; always overwrite so the latest file drives the values
            for entry, value in (
                (self.ps_customer_entry, customer),
                (self.ps_project_entry, project),
                (self.ps_po_entry, po or "TBD"),
                (self.ps_so_entry, so or "TBD"),
            ):
                entry.delete(0, tk.END)
                entry.insert(0, value)

        except Exception as e:
            logging.debug(f"Could not extract metadata from uploaded file: {e}")

    # ------------------------------------------------------------------
    # Packing slip generation
    # ------------------------------------------------------------------

    def generate_packing_slips_from_file(self) -> None:
        """Validate inputs and generate packing slips from the uploaded file."""
        if self.uploaded_file_data is None:
            messagebox.showwarning("No File", "Please upload a file first.")
            return

        customer = self.ps_customer_entry.get().strip()
        project = self.ps_project_entry.get().strip()
        customer_po = self.ps_po_entry.get().strip() or "TBD"
        sales_order = self.ps_so_entry.get().strip() or "TBD"

        if not all([customer, project]):
            messagebox.showerror("Missing Info", "Please fill in Customer and Project fields.")
            return

        self.ps_run_button.config(state=tk.DISABLED)
        self.ps_status_label.config(text="Status: Processing...")
        self.controller.root.update_idletasks()

        # Store form values so the selection dialog can access them
        self._last_customer = customer
        self._last_project = project
        self._last_customer_po = customer_po
        self._last_sales_order = sales_order

        tmp_dir = tempfile.mkdtemp(prefix="LAMIS_")

        try:
            if self._multisheet_device_file:
                processed_data = self._process_multisheet_device_file(self.uploaded_file_path)
            else:
                processed_data = self._process_file_for_packing_slip(self.uploaded_file_data)

            if not processed_data:
                self.ps_status_label.config(text="Status: Ready")
                messagebox.showwarning("No Data", "No valid data found in the file.")
                return

            ip_list = list(processed_data.keys())
            save_path = self.controller.workbook_builder.build_packing_slip_workbook(
                processed_data, ip_list, customer, project, customer_po, sales_order, tmp_dir,
            )

            self.ps_status_label.config(text="Status: Ready")
            out = self.controller.output_screen
            out.insert(tk.END, f"Processing complete — {len(processed_data)} device(s) ready.\n")
            out.see(tk.END)
            self._show_print_selection_dialog(save_path)

        except Exception as e:
            self.ps_status_label.config(text="Status: Error")
            messagebox.showerror("Generation Error", f"Failed to generate packing slips:\n{e}")
            logging.error(f"Packing slip generation error: {e}")
        finally:
            self.ps_run_button.config(state=tk.NORMAL)
            shutil.rmtree(tmp_dir, ignore_errors=True)

    # ------------------------------------------------------------------
    # Print selection
    # ------------------------------------------------------------------

    def _show_print_selection_dialog(self, workbook_path: str) -> None:
        """Open a dialog listing device sheets as checkboxes so the user can
        choose which ones to send to the printer."""
        try:
            wb = openpyxl.load_workbook(workbook_path, read_only=True)
            summary_sheets = [n for n in wb.sheetnames if "summary" in n.lower()]
            device_sheets = [n for n in wb.sheetnames if "summary" not in n.lower()]
            wb.close()
        except Exception as e:
            logging.error(f"Could not read workbook for print selection: {e}")
            return

        if not device_sheets:
            return

        dialog = tk.Toplevel(self.controller.root)
        dialog.title("Select Sheets to Print")
        dialog.grab_set()
        dialog.resizable(False, False)

        tk.Label(dialog, text="Select device sheets to print:", font=("TkDefaultFont", 10, "bold")).pack(
            padx=15, pady=(12, 4), anchor="w"
        )

        # Scrollable checkbox list
        list_frame = ttk.Frame(dialog)
        list_frame.pack(fill=tk.BOTH, expand=True, padx=15, pady=4)

        canvas_height = min(300, len(device_sheets) * 28 + 10)
        canvas = tk.Canvas(list_frame, height=canvas_height, highlightthickness=0)
        scrollbar = ttk.Scrollbar(list_frame, orient="vertical", command=canvas.yview)
        inner = ttk.Frame(canvas)
        inner.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all")),
        )
        canvas.create_window((0, 0), window=inner, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        if len(device_sheets) > 10:
            scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        vars_: Dict[str, tk.BooleanVar] = {}
        for name in device_sheets:
            var = tk.BooleanVar(value=True)
            vars_[name] = var
            ttk.Checkbutton(inner, text=name, variable=var).pack(anchor="w", padx=5, pady=2)

        # Select / Deselect All
        sel_frame = ttk.Frame(dialog)
        sel_frame.pack(fill=tk.X, padx=15, pady=(2, 6))
        tk.Button(sel_frame, text="Select All",   command=lambda: [v.set(True)  for v in vars_.values()]).pack(side=tk.LEFT, padx=2)
        tk.Button(sel_frame, text="Deselect All", command=lambda: [v.set(False) for v in vars_.values()]).pack(side=tk.LEFT, padx=2)

        # Consolidated vs Individual toggle
        ttk.Separator(dialog, orient="horizontal").pack(fill=tk.X, padx=15, pady=(4, 0))
        mode_frame = ttk.LabelFrame(dialog, text="Output Mode")
        mode_frame.pack(fill=tk.X, padx=15, pady=(6, 4))
        print_mode = tk.StringVar(value="consolidated")
        tk.Radiobutton(
            mode_frame, text="Consolidated  (save all selected into one workbook)",
            variable=print_mode, value="consolidated",
        ).pack(anchor="w", padx=10, pady=2)
        tk.Radiobutton(
            mode_frame, text="Individual  (save one workbook per device)",
            variable=print_mode, value="individual",
        ).pack(anchor="w", padx=10, pady=2)

        def on_print() -> None:
            selected = [name for name, var in vars_.items() if var.get()]
            if not selected:
                messagebox.showwarning("No Selection", "Please select at least one sheet.", parent=dialog)
                return
            mode = print_mode.get()
            dialog.destroy()
            self._print_selected_sheets(workbook_path, selected, summary_sheets, mode)

        btn_frame = ttk.Frame(dialog)
        btn_frame.pack(fill=tk.X, padx=15, pady=(4, 12))
        tk.Button(btn_frame, text="Save / Open", command=on_print, width=16).pack(side=tk.RIGHT, padx=2)
        tk.Button(btn_frame, text="Cancel", command=dialog.destroy, width=10).pack(side=tk.RIGHT, padx=2)

        dialog.wait_window()

    def _print_selected_sheets(self, source_path: str, selected_sheets: List[str], summary_sheets: List[str], mode: str = "consolidated") -> None:
        """Save selected sheets to a workbook and open it.

        mode='consolidated': prompt for a single save path; one workbook with
                             the summary + all selected device sheets.
        mode='individual':   prompt for a save folder; one workbook per device,
                             each containing the summary + that device's sheet.
        """
        try:
            base_name = os.path.splitext(os.path.basename(source_path))[0]

            if mode == "individual":
                save_folder = filedialog.askdirectory(title="Select Folder to Save Individual Packing Slips")
                if not save_folder:
                    return  # user cancelled

                saved = 0
                for sheet_name in selected_sheets:
                    wb_src = openpyxl.load_workbook(source_path)
                    sheets_to_keep = set(summary_sheets + [sheet_name])
                    for name in list(wb_src.sheetnames):
                        if name not in sheets_to_keep:
                            del wb_src[name]
                    safe = sheet_name.replace("/", "_").replace("\\", "_")
                    save_path = os.path.join(save_folder, f"{base_name}_{safe}.xlsx")
                    autosize = self.controller.workbook_builder.autosize_sheet_columns
                    for ws_name in wb_src.sheetnames:
                        autosize(wb_src[ws_name])
                    wb_src.save(save_path)
                    os.startfile(save_path)
                    saved += 1

                out = self.controller.output_screen
                out.insert(tk.END, f"Saved {saved} individual packing slip(s) to: {save_folder}\n")
                out.see(tk.END)
                logging.info(f"Individual save: {saved} file(s) → {save_folder}")

            else:  # consolidated — single sheet, all line items from all selected devices
                safe_customer = self._last_customer.replace(" ", "_")
                safe_project = self._last_project.replace(" ", "_")
                save_path = filedialog.asksaveasfilename(
                    title="Save Consolidated Packing Slip",
                    initialfile=f"PackingSlip_{safe_customer}_{safe_project}_Consolidated.xlsx",
                    defaultextension=".xlsx",
                    filetypes=[("Excel files", "*.xlsx")],
                )
                if not save_path:
                    return

                # Build a fresh single-sheet workbook from the consolidated template
                # (data/LAMIS_Consolidated_Packing_Slip.xlsx) which has Device ID in col B
                base_slip_template = self.controller.workbook_builder.packing_slip_template
                consolidated_template = os.path.join(
                    os.path.dirname(base_slip_template),
                    "LAMIS_Consolidated_Packing_Slip.xlsx",
                )
                if not os.path.exists(consolidated_template):
                    consolidated_template = base_slip_template  # fallback

                wb_out = openpyxl.load_workbook(consolidated_template)
                device_sheet = wb_out.active

                # Helper: write to the top-left (master) cell of any merged range,
                # which is the only writable cell in a merge group.
                def write_cell(ws, coord, value):
                    from openpyxl.utils import coordinate_to_tuple
                    r, c = coordinate_to_tuple(coord)
                    for merged in ws.merged_cells.ranges:
                        if r >= merged.min_row and r <= merged.max_row and c >= merged.min_col and c <= merged.max_col:
                            ws.cell(merged.min_row, merged.min_col).value = value
                            return
                    ws[coord] = value

                # Consolidated template layout:
                #   C5 = Customer value, C6 = Project value
                #   Row 14: headers (Device ID | Customer PO | Part Number | Serial Number | Description | Asset Tag)
                #   Row 15+: data rows  — B=Device ID, C=Customer PO, D=Part#, E=Serial#, F=Description
                write_cell(device_sheet, "C5", self._last_customer or "")
                write_cell(device_sheet, "C6", self._last_project or "")

                # Read line items from every selected device sheet and concatenate
                wb_src = openpyxl.load_workbook(source_path, read_only=True)
                row_num = 15
                for sheet_name in selected_sheets:
                    if sheet_name not in wb_src.sheetnames:
                        continue
                    ws = wb_src[sheet_name]
                    for row in ws.iter_rows(min_row=15, values_only=True):
                        if len(row) < 6:
                            continue
                        _so, po, part, serial, desc = row[1], row[2], row[3], row[4], row[5]
                        has_data = any(
                            str(v).strip() not in ("", "None", "nan")
                            for v in (part, serial, desc)
                            if v is not None
                        )
                        if has_data:
                            device_sheet[f"B{row_num}"] = sheet_name   # Device ID
                            device_sheet[f"C{row_num}"] = po or ""
                            device_sheet[f"D{row_num}"] = part or ""
                            device_sheet[f"E{row_num}"] = serial or ""
                            device_sheet[f"F{row_num}"] = desc or ""
                            row_num += 1
                wb_src.close()

                self.controller.workbook_builder.autosize_sheet_columns(device_sheet)
                wb_out.save(save_path)
                os.startfile(save_path)

                out = self.controller.output_screen
                out.insert(tk.END, f"Consolidated packing slip saved: {save_path}\n")
                out.see(tk.END)
                logging.info(f"Consolidated save: {row_num - 15} line item(s) → {save_path}")

        except Exception as e:
            messagebox.showerror("Print Error", f"Failed to prepare sheets for printing:\n{e}")
            logging.error(f"Print selection error: {e}")

    # ------------------------------------------------------------------
    # File processing
    # ------------------------------------------------------------------

    def _process_multisheet_device_file(self, file_path: str) -> Dict[str, pd.DataFrame]:
        """Read a multi-sheet device-report Excel file where each sheet is one device.

        Auto-detects the header row by scanning for 'PART NUMBER' / 'SERIAL NUMBER'.
        Extracts the source IP from sheet metadata and uses it as the dict key.
        Adds a 'System Name' column (= sheet name) so the workbook builder can
        use the sheet name as the device name.
        """
        processed_data: Dict[str, pd.DataFrame] = {}
        try:
            xl = pd.ExcelFile(file_path)
            for sheet_name in xl.sheet_names:
                df_raw = pd.read_excel(file_path, sheet_name=sheet_name, header=None)

                # Extract source IP from metadata (typically row 4, column 5)
                ip_address = sheet_name  # fallback to sheet name
                try:
                    val = str(df_raw.iloc[4, 5]).strip()
                    if val and val.lower() != "nan":
                        ip_address = val
                except (IndexError, KeyError):
                    pass

                # Ensure key uniqueness if multiple devices share the same IP
                if ip_address in processed_data:
                    ip_address = f"{ip_address}_{sheet_name}"

                # Find the header row: first row containing 'PART NUMBER' or 'SERIAL NUMBER'
                header_row = None
                for idx, row in df_raw.iterrows():
                    row_vals = [str(v).upper() for v in row if str(v).lower() != "nan"]
                    joined = " ".join(row_vals)
                    if "PART NUMBER" in joined or "SERIAL NUMBER" in joined:
                        header_row = idx
                        break

                if header_row is None:
                    logging.warning(f"Sheet '{sheet_name}': no header row found, skipping")
                    continue

                df = pd.read_excel(file_path, sheet_name=sheet_name, header=header_row)

                # Drop rows where all key columns are empty
                relevant_cols = [
                    c for c in df.columns
                    if any(k in str(c).upper() for k in ("PART NUMBER", "SERIAL NUMBER", "DESCRIPTION"))
                ]
                if relevant_cols:
                    df = df.dropna(subset=relevant_cols, how="all")
                    df = df[
                        ~df[relevant_cols].apply(
                            lambda r: all(str(v).strip() in ("", "nan") for v in r), axis=1
                        )
                    ]

                # Add System Name column so workbook builder uses sheet name as device name
                df.insert(0, "System Name", sheet_name)

                if not df.empty:
                    processed_data[ip_address] = df.reset_index(drop=True)

            return processed_data

        except Exception as e:
            logging.error(f"Error processing multi-sheet device file: {e}")
            raise

    def _process_file_for_packing_slip(self, df: pd.DataFrame) -> Dict[str, pd.DataFrame]:
        """Convert uploaded DataFrame to per-device dict for the workbook builder."""
        processed_data: Dict[str, pd.DataFrame] = {}
        try:
            # Use lowercased column names only for device key detection; preserve
            # original column casing so the workbook builder can access "Part Number" etc.
            col_lower_map = {col: str(col).lower() for col in df.columns}

            # Check patterns from most specific to least specific to avoid
            # matching unrelated columns like "Part Name" or "File Name".
            # "ip" is NOT used as a bare substring because it matches "Description".
            device_key = None
            priority_patterns = ["system name", "device", "ip address", " ip"]
            fallback_patterns = ["name"]
            for pattern_list in (priority_patterns, fallback_patterns):
                for pattern in pattern_list:
                    for col, col_lower in col_lower_map.items():
                        if pattern in col_lower or (col_lower.strip() == "ip" and pattern == " ip"):
                            device_key = col
                            break
                    if device_key:
                        break
                if device_key:
                    break

            if device_key:
                for device_id, group in df.groupby(device_key, sort=False):
                    processed_data[str(device_id)] = group.reset_index(drop=True)
            else:
                processed_data["Device_0"] = df.reset_index(drop=True)

            return processed_data

        except Exception as e:
            logging.error(f"Error processing file: {e}")
            raise
