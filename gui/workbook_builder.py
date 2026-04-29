"""
WorkbookBuilder: Excel workbook generation logic for ATLAS.

Separated from the main GUI class so it can be tested and maintained independently.
Handles both inventory reports and packing slip workbooks.
"""

from datetime import datetime
from copy import copy
import logging
import os
import re
import shutil
import sqlite3
import tempfile
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
import pandas as pd

from utils.helpers import extract_ip_sort_key


class WorkbookBuilder:
    """Builds Excel workbooks for inventory reports and packing slips."""

    def __init__(self, db_cache: Any, template_path: str, packing_slip_template: str) -> None:
        self.db_cache = db_cache
        self.template_path = template_path
        self.packing_slip_template = packing_slip_template
        self._export_running = False

    # ------------------------------------------------------------------
    # Security helpers
    # ------------------------------------------------------------------

    @staticmethod
    def _sanitize_cell(value: Any) -> Any:
        """Prevent Excel formula injection AND strip illegal control chars.

        - Strips NUL/control characters that openpyxl rejects with
          ``IllegalCharacterError`` (these can leak in from raw device
          telnet output, e.g. the 1830 sends ``\\x0d\\x00`` sequences).
        - Prefixes formula-like leading characters (=, +, -, @, tab, CR)
          with a single quote so Excel treats them as text.
        """
        if isinstance(value, str):
            if value:
                # openpyxl rejects \x00-\x08, \x0B, \x0C, \x0E-\x1F
                value = re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f]", "", value)
            if value and value[0] in ('=', '+', '-', '@', '\t', '\r'):
                return "'" + value
        return value

    # ------------------------------------------------------------------
    # Sheet utilities
    # ------------------------------------------------------------------

    def autosize_sheet_columns(self, sheet: Any, min_width: int = 10, max_width: int = 60) -> None:
        """Auto-size all populated columns in a worksheet with sane bounds."""
        max_col = sheet.max_column or 0
        max_row = sheet.max_row or 0
        if max_col <= 0 or max_row <= 0:
            return

        for col_idx in range(1, max_col + 1):
            longest = 0
            for row_idx in range(1, max_row + 1):
                val = sheet.cell(row=row_idx, column=col_idx).value
                if val is None:
                    continue
                text_len = len(str(val))
                if text_len > longest:
                    longest = text_len

            if longest == 0:
                continue

            width = min(max(longest + 2, min_width), max_width)
            sheet.column_dimensions[get_column_letter(col_idx)].width = width

    def autosize_workbook_columns(self, wb: Any, min_width: int = 10, max_width: int = 60) -> None:
        """Auto-size populated columns in every visible sheet of a workbook."""
        for sheet_name in wb.sheetnames:
            try:
                self.autosize_sheet_columns(wb[sheet_name], min_width, max_width)
            except Exception as exc:
                logging.warning(f"Autosize failed for sheet '{sheet_name}': {exc}")

    def copy_sheet(self, source_sheet: Any, target_wb: Any, new_sheet_name: str) -> Any:
        """Copy an entire sheet from one workbook to another while preserving formatting."""
        new_sheet = target_wb.create_sheet(title=new_sheet_name)

        for row in source_sheet.iter_rows():
            for cell in row:
                new_sheet[cell.coordinate].value = cell.value
                if cell.has_style:
                    # Openpyxl style objects must be cloned when copied across workbooks.
                    new_sheet[cell.coordinate].font = copy(cell.font)
                    new_sheet[cell.coordinate].border = copy(cell.border)
                    new_sheet[cell.coordinate].fill = copy(cell.fill)
                    new_sheet[cell.coordinate].number_format = cell.number_format
                    new_sheet[cell.coordinate].protection = copy(cell.protection)
                    new_sheet[cell.coordinate].alignment = copy(cell.alignment)

        for key, dimension in source_sheet.column_dimensions.items():
            new_sheet.column_dimensions[key].width = dimension.width
            new_sheet.column_dimensions[key].hidden = dimension.hidden

        for key, dimension in source_sheet.row_dimensions.items():
            new_sheet.row_dimensions[key].height = dimension.height
            new_sheet.row_dimensions[key].hidden = dimension.hidden

        for merged_range in source_sheet.merged_cells.ranges:
            new_sheet.merge_cells(str(merged_range))

        if source_sheet.freeze_panes:
            new_sheet.freeze_panes = source_sheet.freeze_panes

        return new_sheet

    # ------------------------------------------------------------------
    # Shared summary-sheet helpers (used by both workbook builders)
    # ------------------------------------------------------------------

    def _setup_summary_sheet_header(self, summary_sheet: Any, customer: str, project: str, ip_list: List[str]) -> None:
        """Common header setup for summary sheets in both report and packing slip workbooks."""
        summary_sheet["B5"] = "Customer"
        summary_sheet["B7"] = customer or ""
        summary_sheet["D5"] = "Project"
        summary_sheet["D7"] = project or ""

        capture_time = datetime.now().strftime('%Y-%m-%d @ %H:%M:%S')
        summary_sheet["A2"] = f"Capture Time = {capture_time}"
        summary_sheet["A2"].font = Font(color="00FF00", bold=True)
        summary_sheet["A2"].fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")

        summary_sheet["B9"] = "#"
        summary_sheet["C9"] = "IP Address"
        summary_sheet["D9"] = "Device Name"

        summary_sheet["F5"] = "Device Count"
        summary_sheet["F7"] = 0  # Updated with actual count after processing

    def _format_summary_sheet(self, summary_sheet: Any, data_row_count: int) -> None:
        """Apply a consistent visual layout to summary sheets."""
        accent_fill = PatternFill(start_color="FF156082", end_color="FF156082", fill_type="solid")
        dark_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
        white_font = Font(color="FFFFFFFF", bold=True)
        bold_font = Font(bold=True)
        thin_side = Side(style="thin", color="FF156082")
        border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        center = Alignment(horizontal="center", vertical="center")

        summary_sheet.merge_cells("B2:F3")
        summary_sheet["B2"] = "Device Summary"
        summary_sheet["B2"].fill = accent_fill
        summary_sheet["B2"].font = Font(color="FFFFFFFF", bold=True, size=18)
        summary_sheet["B2"].alignment = center

        for row in range(5, 8):
            for col in range(2, 7):
                cell = summary_sheet.cell(row=row, column=col)
                cell.border = border
                cell.alignment = center if row == 5 else Alignment(vertical="center")
                if row == 5:
                    cell.fill = accent_fill
                    cell.font = white_font

        summary_sheet["A2"].fill = dark_fill

        for col in range(2, 5):
            cell = summary_sheet.cell(row=9, column=col)
            cell.fill = accent_fill
            cell.font = white_font
            cell.border = border
            cell.alignment = center

        end_row = max(10, 9 + data_row_count)
        for row in range(10, end_row + 1):
            for col in range(2, 5):
                cell = summary_sheet.cell(row=row, column=col)
                cell.border = border
                if col == 2:
                    cell.alignment = center

        summary_sheet.freeze_panes = "B10"

    def _format_packing_slip_summary(self, summary_sheet: Any, data_row_count: int) -> None:
        """Apply the same visual layout as the inventory summary to packing slip summary sheets."""
        accent_fill = PatternFill(start_color="FF156082", end_color="FF156082", fill_type="solid")
        dark_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
        white_font = Font(color="FFFFFFFF", bold=True)
        thin_side = Side(style="thin", color="FF156082")
        border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        center = Alignment(horizontal="center", vertical="center")

        # Title block
        summary_sheet.merge_cells("B2:H3")
        summary_sheet["B2"] = "Packing Slip Summary"
        summary_sheet["B2"].fill = accent_fill
        summary_sheet["B2"].font = Font(color="FFFFFFFF", bold=True, size=18)
        summary_sheet["B2"].alignment = center
        summary_sheet["A2"].fill = dark_fill

        # Header labels and styling at row 5
        headers = {2: "Customer", 4: "Project", 6: "IP Address", 8: "Device Name"}
        for col, label in headers.items():
            cell = summary_sheet.cell(row=5, column=col)
            cell.value = label
            cell.fill = accent_fill
            cell.font = white_font
            cell.border = border
            cell.alignment = center

        # Device count info block
        summary_sheet["J5"] = "Device Count"
        summary_sheet["J5"].fill = accent_fill
        summary_sheet["J5"].font = white_font
        summary_sheet["J5"].border = border
        summary_sheet["J5"].alignment = center
        summary_sheet["J6"] = data_row_count
        summary_sheet["J6"].border = border
        summary_sheet["J6"].alignment = center

        # Borders on data rows
        end_row = max(7, 6 + data_row_count)
        for row in range(7, end_row + 1):
            for col in [2, 4, 6, 8]:
                summary_sheet.cell(row=row, column=col).border = border

        summary_sheet.freeze_panes = "B7"

    def _populate_summary_table(self, summary_sheet: Any, summary_items: List[tuple], start_row: int = 10) -> None:
        """Populate summary table with device entries, sorted by IP."""
        ordered_items = sorted(summary_items, key=lambda item: extract_ip_sort_key(item[0]))

        for row_offset, (ip, device_name, sheet_title) in enumerate(ordered_items):
            row_num = start_row + row_offset
            summary_sheet[f"B{row_num}"] = row_offset + 1
            summary_sheet[f"C{row_num}"] = str(ip)
            summary_sheet[f"D{row_num}"] = str(device_name)
            summary_sheet[f"D{row_num}"].hyperlink = f"#'{sheet_title}'!A1"
            summary_sheet[f"D{row_num}"].style = "Hyperlink"

        summary_sheet["F7"] = len(ordered_items)
        self._format_summary_sheet(summary_sheet, len(ordered_items))
        self.autosize_sheet_columns(summary_sheet)

    # ------------------------------------------------------------------
    # Data combination
    # ------------------------------------------------------------------

    def combine_and_format_data(self, ip_data: Dict[str, pd.DataFrame]) -> pd.DataFrame:
        """Concatenate per-IP DataFrames from *ip_data* into a single DataFrame.

        Accepts dict values that are either a bare DataFrame or a dict with a
        "DataFrame" key.  Logs a warning and skips any entry that doesn't
        match either shape.  Returns an empty DataFrame when *ip_data* is
        empty or contains no usable data.
        """
        all_data = []
        logging.info(f"Starting combination of {len(ip_data)} data entries.")

        for key, value in ip_data.items():
            if isinstance(value, pd.DataFrame):
                all_data.append(value)
                logging.info(f"Processed DataFrame under key '{key}' with {len(value)} rows.")
            elif isinstance(value, dict) and "DataFrame" in value and isinstance(value["DataFrame"], pd.DataFrame):
                df = value["DataFrame"]
                all_data.append(df)
                logging.info(f"Unwrapped DataFrame under key '{key}' with {len(df)} rows.")
            else:
                logging.warning(f"Expected DataFrame under key '{key}' but got {type(value)}. Skipping.")

        if all_data:
            combined_df = pd.concat(all_data, ignore_index=True)
            logging.info(f"Combined DataFrame created with {len(combined_df)} rows from {len(all_data)} DataFrames.")
        else:
            combined_df = pd.DataFrame()
            logging.warning("No data to combine. Returning empty DataFrame.")

        return combined_df

    # ------------------------------------------------------------------
    # Inventory report workbook
    # ------------------------------------------------------------------

    def build_report_workbook(self, outputs: Dict[str, Any], output_file: str, customer: str = "", project: str = "", customer_po: str = "", sales_order: str = "", append_mode: bool = False) -> Dict[str, Any]:
        """Build (or append to) an inventory report Excel workbook.

        Args:
            outputs: Per-IP data keyed by IP address.
            output_file: Destination .xlsx path.
            customer: Customer name written to the summary sheet.
            project: Project name written to the summary sheet.
            customer_po: Purchase order number.
            sales_order: Sales order number.
            append_mode: When True, open an existing workbook and add new
                device sheets rather than creating a new file.

        Returns:
            Dict of per-IP processed data (used when generating packing slips
            immediately after export).

        Raises:
            RuntimeError: If an export is already in progress.
        """
        if getattr(self, "_export_running", False):
            raise RuntimeError("Export already running")
        self._export_running = True

        logging.info("Starting Excel export process...")
        try:
            def clean_str(v: object) -> str:
                if v is None:
                    return ""
                s = str(v).strip()
                return "" if s.lower() == "nan" else s

            def nonempty_col(col):
                if col is None:
                    return None
                s = col.astype(str).str.strip()
                return s.ne("") & s.str.lower().ne("nan")

            def first_nonempty(series, fallback=""):
                if series is None:
                    return fallback
                for v in series:
                    s = clean_str(v)
                    if s:
                        return s
                return fallback

            if not os.path.exists(self.template_path):
                raise FileNotFoundError(f"Template file not found at {self.template_path}")

            if append_mode:
                if not os.path.exists(output_file):
                    raise FileNotFoundError(f"Existing report file not found at {output_file}")
                wb = openpyxl.load_workbook(output_file)
                sheet = None
            else:
                wb = openpyxl.load_workbook(self.template_path)
                sheet = wb.active

            template_wb = openpyxl.load_workbook(self.template_path)
            template_sheet = template_wb.active

            db_exists = os.path.isfile(self.db_cache.db_path)
            logging.info(f"[EXCEL] Using DB cache path: {self.db_cache.db_path} (exists={db_exists})")
            logging.info(f"Starting Excel export for {len(outputs)} devices.")

            processed_data = {}
            summary_index = {}

            if "Summary" in wb.sheetnames:
                summary_sheet = wb["Summary"]
            else:
                summary_sheet = wb.create_sheet(title="Summary", index=0)

            if append_mode:
                for row_num in range(10, summary_sheet.max_row + 1):
                    existing_ip = summary_sheet[f"C{row_num}"].value
                    existing_name = summary_sheet[f"D{row_num}"].value
                    if existing_ip is None or existing_name is None:
                        continue

                    existing_ip_str = str(existing_ip).strip()
                    existing_name_str = str(existing_name).strip()
                    if not existing_ip_str:
                        continue

                    sheet_title = None
                    if summary_sheet[f"D{row_num}"].hyperlink and summary_sheet[f"D{row_num}"].hyperlink.target:
                        target = summary_sheet[f"D{row_num}"].hyperlink.target
                        match = re.match(r"#'(.+)'!", str(target))
                        if match:
                            sheet_title = match.group(1)

                    if sheet_title and sheet_title in wb.sheetnames:
                        summary_index[existing_ip_str] = (existing_name_str, sheet_title)

            self._setup_summary_sheet_header(summary_sheet, customer, project, list(outputs.keys()))

            def make_unique_sheet_title(base_name):
                clean_base = re.sub(r'[^a-zA-Z0-9_]', '_', str(base_name).strip())[:31]
                if not clean_base:
                    clean_base = "Device"

                if clean_base not in wb.sheetnames:
                    return clean_base

                counter = 2
                while True:
                    suffix = f"_{counter}"
                    candidate = f"{clean_base[:31 - len(suffix)]}{suffix}"
                    if candidate not in wb.sheetnames:
                        return candidate
                    counter += 1

            ordered_outputs = sorted(outputs.items(), key=lambda item: extract_ip_sort_key(item[0]))
            for seq, (ip, data_dict) in enumerate(ordered_outputs, start=1):
                new_sheet = None
                try:
                    logging.info(f"Processing data for IP {ip}.")
                    combined_df = self.combine_and_format_data(data_dict)
                    if combined_df.empty:
                        logging.warning(f"No data to write for IP {ip}")
                        continue

                    system_name = clean_str(first_nonempty(
                        combined_df.get("System Name"),
                        f"System_{ip.replace('.', '_')}"
                    ))[:31].replace(":", "_").replace("/", "_")

                    system_type = clean_str(first_nonempty(combined_df.get("System Type"), ""))
                    if not system_type:
                        system_type = clean_str(first_nonempty(combined_df.get("Type"), ""))
                    system_type = system_type[:31].replace(":", "_").replace("/", "_")

                    source_val = clean_str(first_nonempty(combined_df.get("Source"), ""))

                    logging.info(f"Creating sheet for system '{system_name}' with {len(combined_df)} rows.")
                    ip_key = str(ip)
                    prior_sheet_title = None
                    if append_mode and ip_key in summary_index:
                        prior_sheet_title = summary_index[ip_key][1]

                    new_sheet_title = make_unique_sheet_title(system_name)
                    if append_mode:
                        new_sheet = self.copy_sheet(template_sheet, wb, new_sheet_title)
                    else:
                        new_sheet = wb.copy_worksheet(sheet)
                        new_sheet.title = new_sheet_title

                    # Add quick navigation back to summary for easier review workflow.
                    new_sheet["A1"] = "Return"
                    new_sheet["A1"].hyperlink = f"#'{summary_sheet.title}'!A1"
                    new_sheet["A1"].style = "Hyperlink"

                    new_sheet["C5"] = customer
                    new_sheet["C6"] = project
                    new_sheet["C7"] = customer_po
                    new_sheet["D7"] = sales_order
                    new_sheet["F5"] = source_val
                    new_sheet["F6"] = self._sanitize_cell(system_name)
                    new_sheet["F7"] = self._sanitize_cell(system_type)

                    has_part = combined_df.get("Part Number")
                    has_model = combined_df.get("Model Number")

                    mask_part = nonempty_col(has_part)
                    mask_model = nonempty_col(has_model)

                    if mask_part is not None and mask_model is not None:
                        write_df = combined_df[mask_part | mask_model].copy()
                    elif mask_part is not None:
                        write_df = combined_df[mask_part].copy()
                    elif mask_model is not None:
                        write_df = combined_df[mask_model].copy()
                    else:
                        write_df = combined_df.iloc[0:0].copy()

                    start_row = 15
                    for i, row in write_df.reset_index(drop=True).iterrows():
                        row_num = start_row + i

                        part_number = clean_str(row.get("Part Number", "")) or clean_str(row.get("Model Number", ""))
                        info_type = clean_str(row.get("Information Type", "")).lower()
                        name_value = clean_str(row.get("Name", ""))

                        if "mda card" in info_type:
                            match = re.search(r"\d+", name_value)
                            mda_number = match.group() if match else "Unknown"
                            name_value = f"MDA {mda_number}"

                        type_value = clean_str(row.get("Type", ""))
                        serial_val = clean_str(row.get("Serial Number", ""))

                        if not any([name_value, type_value, part_number, serial_val]):
                            continue

                        description = ""
                        if part_number and db_exists:
                            description = self.db_cache.lookup_part(part_number[:10])
                            if not description or description == "Not Found":
                                try:
                                    with sqlite3.connect(self.db_cache.db_path) as tmpconn:
                                        tcur = tmpconn.cursor()
                                        tcur.execute(
                                            "SELECT description FROM parts WHERE part_number LIKE ?",
                                            (part_number[:10] + "%",),
                                        )
                                        res = tcur.fetchone()
                                        if res:
                                            description = res[0]
                                except Exception as exc:
                                    logging.debug(f"[EXCEL] Fallback DB lookup failed: {exc}")

                        new_sheet[f"B{row_num}"] = self._sanitize_cell(name_value)
                        new_sheet[f"C{row_num}"] = self._sanitize_cell(type_value)
                        new_sheet[f"D{row_num}"] = self._sanitize_cell(part_number)
                        new_sheet[f"E{row_num}"] = self._sanitize_cell(serial_val)
                        new_sheet[f"F{row_num}"] = self._sanitize_cell(description)
                        write_df.at[row.name, "Description"] = description

                    self.autosize_sheet_columns(new_sheet)

                    if prior_sheet_title and prior_sheet_title in wb.sheetnames and prior_sheet_title != summary_sheet.title:
                        del wb[prior_sheet_title]

                    if prior_sheet_title and new_sheet.title != system_name and system_name not in wb.sheetnames:
                        new_sheet.title = system_name

                    processed_data[ip] = write_df
                    summary_index[str(ip)] = (system_name, new_sheet.title)
                except Exception as exc:
                    logging.error(f"Failed to process data for IP {ip}. Error: {exc}")
                    if new_sheet is not None and new_sheet.title in wb.sheetnames:
                        del wb[new_sheet.title]

            if summary_sheet.max_row >= 10:
                for row_num in range(10, summary_sheet.max_row + 1):
                    summary_sheet[f"B{row_num}"] = None
                    summary_sheet[f"C{row_num}"] = None
                    summary_sheet[f"D{row_num}"] = None

            # Populate summary table using shared helper (also sets the correct device count).
            summary_items = [(ip, device_name, sheet_title) for ip, (device_name, sheet_title) in summary_index.items()]
            self._populate_summary_table(summary_sheet, summary_items, start_row=10)

            if not append_mode and sheet and sheet.title in wb.sheetnames:
                wb.remove(sheet)

            # Keep tabs ordered by IP sequence, with Summary first.
            ordered_summary_items = sorted(summary_items, key=lambda item: extract_ip_sort_key(item[0]))
            ordered_device_tabs = [sheet_title for (_, _, sheet_title) in ordered_summary_items if sheet_title in wb.sheetnames]
            pinned_tabs = [summary_sheet.title] + ordered_device_tabs
            remaining_tabs = [name for name in wb.sheetnames if name not in pinned_tabs]
            ordered_tabs = pinned_tabs + remaining_tabs
            wb._sheets = [wb[name] for name in ordered_tabs]

            save_dir = os.path.dirname(output_file)
            os.makedirs(save_dir, exist_ok=True)
            self.autosize_workbook_columns(wb)
            wb.save(output_file)
            logging.info(f"Data successfully saved to {output_file}")
            template_wb.close()
            return processed_data
        finally:
            self._export_running = False

    # ------------------------------------------------------------------
    # Packing slip workbook
    # ------------------------------------------------------------------

    # ------------------------------------------------------------------
    # Nokia PSI report workbook
    # ------------------------------------------------------------------

    def build_psi_report_workbook(
        self,
        outputs: Dict[str, Any],
        output_file: str,
        customer: str = "",
        project: str = "",
        customer_po: str = "",
        sales_order: str = "",
        append_mode: bool = False,
        psi_template_path: Optional[str] = None,
    ) -> Dict[str, Any]:
        """Build a Nokia PSI inventory report using the PSI-specific template.

        Template row layout
        -------------------
        15  - 54  : Equipment inventory (shelf / card / module)
        62  - 71  : Software information
        74  - 103 : Slot status
        106 - 115 : Redundancy information
        118 - 127 : Power feed status
        130 - 169 : Interface topology
        """
        if psi_template_path is None:
            psi_template_path = os.path.join(
                os.path.dirname(self.template_path), "Nokia_PSI_Report_Template.xlsx"
            )
        if not os.path.exists(psi_template_path):
            raise FileNotFoundError(f"PSI template not found at {psi_template_path}")

        template_wb = openpyxl.load_workbook(psi_template_path)
        template_sheet = template_wb.active

        if append_mode and os.path.exists(output_file):
            wb = openpyxl.load_workbook(output_file)
            base_sheet = None
        else:
            wb = openpyxl.load_workbook(psi_template_path)
            base_sheet = wb.active

        summary_sheet = (
            wb["Summary"] if "Summary" in wb.sheetnames
            else wb.create_sheet(title="Summary", index=0)
        )
        self._setup_summary_sheet_header(summary_sheet, customer, project, list(outputs.keys()))

        def _s(v):
            if v is None:
                return ""
            s = str(v).strip()
            return "" if s.lower() == "nan" else s

        def _first(series, fallback=""):
            if series is None:
                return fallback
            for v in series:
                s = _s(v)
                if s:
                    return s
            return fallback

        def _df(entry) -> pd.DataFrame:
            if isinstance(entry, dict) and "DataFrame" in entry:
                return entry["DataFrame"]
            if isinstance(entry, pd.DataFrame):
                return entry
            return pd.DataFrame()

        sanitize = self._sanitize_cell

        def _wr(ws, r, b="", c="", d="", e="", f=""):
            ws[f"B{r}"] = sanitize(b)
            ws[f"C{r}"] = sanitize(c)
            ws[f"D{r}"] = sanitize(d)
            ws[f"E{r}"] = sanitize(e)
            ws[f"F{r}"] = sanitize(f)

        def _write_inventory(ws, df, start=15, end=54):
            for i, (_, row) in enumerate(df.iterrows()):
                r = start + i
                if r > end:
                    break
                _wr(ws, r,
                    b=_s(row.get("Name", "")),
                    c=_s(row.get("Type", "")),
                    d=_s(row.get("Part Number", "")),
                    e=_s(row.get("Serial Number", "")),
                    f=_s(row.get("Description", "")),
                )

        def _write_software(ws, df, start=62, end=71):
            for i, (_, row) in enumerate(df.iterrows()):
                r = start + i
                if r > end:
                    break
                desc = _s(row.get("Description", ""))
                m = re.search(r"RPMS Loaded:\s*(\S+)\s*/\s*(\S+)", desc)
                _wr(ws, r,
                    b=_s(row.get("Name", "")),
                    c=_s(row.get("Part Number", "")),
                    d=m.group(1) if m else "",
                    e=m.group(2) if m else "",
                )

        def _write_slot(ws, df, start=74, end=103):
            for i, (_, row) in enumerate(df.iterrows()):
                r = start + i
                if r > end:
                    break
                desc = _s(row.get("Description", ""))
                m = re.search(r"Admin:\s*(\S+)\s*\|\s*Oper:\s*(\S+)(?:\s*\|\s*(.+))?", desc)
                admin_state = m.group(1) if m else ""
                oper_state = m.group(2) if m else ""
                qualifier = m.group(3).strip() if m and m.group(3) else ""
                _wr(ws, r,
                    b=_s(row.get("Name", "")),
                    c=_s(row.get("Part Number", "")),
                    d=_s(row.get("Present Type", "")),
                    e=admin_state,
                    f=f"{oper_state}{' | ' + qualifier if qualifier else ''}",
                )

        def _write_redundancy(ws, df, start=106, end=115):
            for i, (_, row) in enumerate(df.iterrows()):
                r = start + i
                if r > end:
                    break
                desc = _s(row.get("Description", ""))
                m_c = re.search(r"Clock Switch:\s*(\S+)", desc)
                m_e = re.search(r"EC Selection:\s*(\S+)", desc)
                _wr(ws, r,
                    b=_s(row.get("Name", "")),
                    c=m_c.group(1) if m_c else "",
                    d=m_e.group(1) if m_e else "",
                )

        def _write_power(ws, df, start=118, end=127):
            for i, (_, row) in enumerate(df.iterrows()):
                r = start + i
                if r > end:
                    break
                desc = _s(row.get("Description", ""))
                m_a = re.search(r"Admin:\s*(\S+)", desc)
                m_o = re.search(r"Oper:\s*(\S+)", desc)
                _wr(ws, r,
                    b=_s(row.get("Name", "")),
                    c=_s(row.get("Type", "")),
                    d=m_a.group(1) if m_a else "",
                    e=m_o.group(1) if m_o else "",
                )

        def _write_topology(ws, df, start=130, end=169):
            for i, (_, row) in enumerate(df.iterrows()):
                r = start + i
                if r > end:
                    break
                desc = _s(row.get("Description", ""))
                m_c = re.search(r"Connected To:\s*(\S+)", desc)
                m_f = re.search(r"From:\s*(\S+)", desc)
                _wr(ws, r,
                    b=_s(row.get("Name", "")),
                    c=_s(row.get("Type", "")),
                    d=m_c.group(1) if m_c else "",
                    e=m_f.group(1) if m_f else "",
                )

        processed_data: Dict[str, Any] = {}
        summary_index: Dict[str, tuple] = {}

        def _unique_title(base):
            clean = re.sub(r"[^a-zA-Z0-9_]", "_", str(base).strip())[:31] or "Device"
            if clean not in wb.sheetnames:
                return clean
            n = 2
            while True:
                suf = f"_{n}"
                cand = f"{clean[:31 - len(suf)]}{suf}"
                if cand not in wb.sheetnames:
                    return cand
                n += 1

        for ip, data_dict in sorted(outputs.items(), key=lambda x: extract_ip_sort_key(x[0])):
            try:
                system_name = system_type = source_val = ""
                for key in ("shelf_detail", "shelf_inventory", "card_inventory"):
                    entry = data_dict.get(key)
                    if entry:
                        df = _df(entry)
                        if not df.empty:
                            if not system_name:
                                system_name = _s(_first(df.get("System Name")))
                            if not system_type:
                                system_type = _s(_first(df.get("System Type")))
                            if not source_val:
                                source_val = _s(_first(df.get("Source"), str(ip)))
                    if system_name:
                        break
                if not system_name:
                    system_name = f"PSI_{ip.replace('.', '_')}"
                if not source_val:
                    source_val = str(ip)

                title = _unique_title(system_name)
                if append_mode:
                    ns = self.copy_sheet(template_sheet, wb, title)
                else:
                    ns = wb.copy_worksheet(base_sheet)
                    ns.title = title

                ns["A1"] = "Return"
                ns["A1"].hyperlink = f"#'{summary_sheet.title}'!A1"
                ns["A1"].style = "Hyperlink"
                ns["C5"] = customer
                ns["C6"] = project
                ns["C7"] = customer_po
                ns["D7"] = sales_order
                ns["F5"] = source_val
                ns["F6"] = self._sanitize_cell(system_name)
                ns["F7"] = self._sanitize_cell(system_type)

                shelf_detail_df = _df(data_dict.get("shelf_detail"))
                shelf_inventory_df = _df(data_dict.get("shelf_inventory")).copy()
                card_inventory_df = _df(data_dict.get("card_inventory"))
                module_inventory_df = _df(data_dict.get("module_inventory"))

                inv_parts = []
                if not shelf_inventory_df.empty:
                    if not shelf_detail_df.empty:
                        main_shelf_name = _s(_first(shelf_detail_df.get("Name"), "Main Shelf"))
                        mask = shelf_inventory_df.get("Type", pd.Series(dtype=str)).astype(str).str.strip().eq("Shelf")
                        if mask.any():
                            shelf_row_index = shelf_inventory_df[mask].index[0]
                            shelf_inventory_df.at[shelf_row_index, "Name"] = main_shelf_name
                    inv_parts.append(shelf_inventory_df)
                elif not shelf_detail_df.empty:
                    inv_parts.append(shelf_detail_df)

                for df in (card_inventory_df, module_inventory_df):
                    if not df.empty:
                        inv_parts.append(df)

                if inv_parts:
                    _write_inventory(ns, pd.concat(inv_parts, ignore_index=True))

                _write_software(ns, _df(data_dict.get("software_info")))
                _write_slot(ns, _df(data_dict.get("slot_info")))
                _write_redundancy(ns, _df(data_dict.get("redundancy_info")))
                _write_power(ns, _df(data_dict.get("power_info")))
                _write_topology(ns, _df(data_dict.get("topology")))

                self.autosize_sheet_columns(ns)
                summary_index[str(ip)] = (system_name, title)
                processed_data[ip] = pd.DataFrame()

            except Exception as exc:
                logging.error(f"PSI report: failed for IP {ip}: {exc}", exc_info=True)

        items = [(ip, n, t) for ip, (n, t) in summary_index.items()]
        self._populate_summary_table(summary_sheet, items, start_row=10)

        if not append_mode and base_sheet and base_sheet.title in wb.sheetnames and len(wb.sheetnames) > 1:
            wb.remove(base_sheet)

        ordered = [summary_sheet.title] + [
            t for _, _, t in sorted(items, key=lambda x: extract_ip_sort_key(x[0]))
            if t in wb.sheetnames
        ]
        remaining = [n for n in wb.sheetnames if n not in ordered]
        wb._sheets = [wb[n] for n in ordered + remaining]

        save_dir = os.path.dirname(output_file)
        if save_dir:
            os.makedirs(save_dir, exist_ok=True)
        self.autosize_workbook_columns(wb)
        wb.save(output_file)
        logging.info(f"PSI report saved: {output_file}")
        template_wb.close()
        return processed_data

    # ------------------------------------------------------------------
    # Unified multi-template workbook
    # ------------------------------------------------------------------

    def build_unified_report_workbook(
        self,
        family_buckets: Dict[str, Dict[str, Any]],
        output_file: str,
        customer: str = "",
        project: str = "",
        customer_po: str = "",
        sales_order: str = "",
        append_mode: bool = False,
        rls_template_path: Optional[str] = None,
        psi_template_path: Optional[str] = None,
    ) -> Dict[str, Any]:
        """Build a single workbook containing per-device sheets from multiple
        template families.

        ``family_buckets`` maps a family key (``"rls"``, ``"psi"``,
        ``"default"``) to ``{ip: device_outputs}``. Each non-empty bucket is
        first rendered to a temporary workbook by the appropriate
        family-specific builder, then their device sheets are merged into a
        single output workbook with one unified Summary sheet.
        """
        # 1. Render each family to a temp file using the existing builders.
        non_empty: Dict[str, Dict[str, Any]] = {
            fam: ips for fam, ips in family_buckets.items() if ips
        }
        if not non_empty:
            logging.warning("build_unified_report_workbook: no devices to write.")
            return {}

        common = dict(
            customer=customer,
            project=project,
            customer_po=customer_po,
            sales_order=sales_order,
            append_mode=False,  # always build temp wbs from scratch
        )

        tempdir = tempfile.mkdtemp(prefix="atlas_unified_")
        family_files: Dict[str, str] = {}
        processed_all: Dict[str, Any] = {}

        try:
            for fam, ips in non_empty.items():
                tmp_path = os.path.join(tempdir, f"{fam}.xlsx")
                if fam == "rls":
                    pd_part = self.build_psi_report_workbook(
                        ips, tmp_path,
                        psi_template_path=rls_template_path or psi_template_path,
                        **common,
                    )
                elif fam == "psi":
                    pd_part = self.build_psi_report_workbook(
                        ips, tmp_path,
                        psi_template_path=psi_template_path,
                        **common,
                    )
                else:
                    pd_part = self.build_report_workbook(ips, tmp_path, **common)
                processed_all.update(pd_part or {})
                family_files[fam] = tmp_path

            # 2. Pick a base workbook to merge into. Order: default, psi, rls
            # so the chrome (theme/styles) of the standard report wins when
            # available — matches existing single-family behavior.
            order = [f for f in ("default", "psi", "rls") if f in family_files]
            base_fam = order[0]
            base_path = family_files[base_fam]

            # If append_mode and the output file already exists, load it as
            # the base instead so we extend the existing report.
            if append_mode and os.path.exists(output_file):
                wb = openpyxl.load_workbook(output_file)
            else:
                wb = openpyxl.load_workbook(base_path)

            # Collect (ip, device_name, sheet_title) entries from each family
            # by scanning each family wb's Summary sheet.
            unified_summary: List[Tuple[str, str, str]] = []

            def _scan_family_summary(src_wb) -> List[Tuple[str, str, str]]:
                items: List[Tuple[str, str, str]] = []
                if "Summary" not in src_wb.sheetnames:
                    return items
                ss = src_wb["Summary"]
                for r in range(10, ss.max_row + 1):
                    ip_val = ss[f"C{r}"].value
                    name_val = ss[f"D{r}"].value
                    if ip_val is None or name_val is None:
                        continue
                    ip_s = str(ip_val).strip()
                    name_s = str(name_val).strip()
                    if not ip_s:
                        continue
                    target = ""
                    hl = ss[f"D{r}"].hyperlink
                    if hl and hl.target:
                        m = re.match(r"#'(.+?)'!", str(hl.target))
                        if m:
                            target = m.group(1)
                    if not target:
                        target = name_s[:31]
                    items.append((ip_s, name_s, target))
                return items

            # Seed unified summary with whatever's already in the base wb.
            unified_summary.extend(_scan_family_summary(wb))

            existing_titles = set(wb.sheetnames)

            def _unique_title(base):
                clean = re.sub(r"[^a-zA-Z0-9_]", "_", str(base).strip())[:31] or "Device"
                if clean not in existing_titles:
                    return clean
                n = 2
                while True:
                    suf = f"_{n}"
                    cand = f"{clean[:31 - len(suf)]}{suf}"
                    if cand not in existing_titles:
                        return cand
                    n += 1

            # 3. Merge sheets from the other family workbooks into wb.
            for fam in order[1:]:
                src_path = family_files[fam]
                src_wb = openpyxl.load_workbook(src_path)
                try:
                    src_items = _scan_family_summary(src_wb)
                    for ip_s, name_s, src_title in src_items:
                        if src_title not in src_wb.sheetnames:
                            logging.warning(
                                f"Unified merge: source sheet '{src_title}' missing in {fam} wb; skipping IP {ip_s}."
                            )
                            continue
                        new_title = _unique_title(src_title)
                        new_sheet = self.copy_sheet(src_wb[src_title], wb, new_title)
                        existing_titles.add(new_title)
                        # Repoint the per-sheet "Return" hyperlink to the
                        # unified Summary (it already points to 'Summary'
                        # which exists in the base, so no change needed —
                        # but enforce it defensively in case styles vary).
                        try:
                            if new_sheet["A1"].value == "Return":
                                new_sheet["A1"].hyperlink = "#'Summary'!A1"
                        except Exception:
                            pass
                        unified_summary.append((ip_s, name_s, new_title))
                finally:
                    src_wb.close()

            # 4. Rebuild Summary in the unified wb covering all devices.
            if "Summary" in wb.sheetnames:
                summary_sheet = wb["Summary"]
                # Clear prior table rows (10+) so we don't double-populate.
                for r in range(10, summary_sheet.max_row + 1):
                    for col in ("B", "C", "D"):
                        cell = summary_sheet[f"{col}{r}"]
                        cell.value = None
                        cell.hyperlink = None
            else:
                summary_sheet = wb.create_sheet(title="Summary", index=0)

            # Move Summary to front.
            if wb.sheetnames[0] != "Summary":
                wb._sheets.insert(0, wb._sheets.pop(wb.sheetnames.index("Summary")))

            all_ips = list({ip for ip in (it[0] for it in unified_summary)})
            self._setup_summary_sheet_header(summary_sheet, customer, project, all_ips)
            self._populate_summary_table(summary_sheet, unified_summary, start_row=10)

            # Order device sheets by IP for predictable layout.
            ordered_titles = [t for _, _, t in sorted(unified_summary, key=lambda x: extract_ip_sort_key(x[0])) if t in wb.sheetnames]
            remaining = [n for n in wb.sheetnames if n != "Summary" and n not in ordered_titles]
            wb._sheets = [wb["Summary"]] + [wb[t] for t in ordered_titles] + [wb[n] for n in remaining]

            save_dir = os.path.dirname(output_file)
            if save_dir:
                os.makedirs(save_dir, exist_ok=True)
            self.autosize_workbook_columns(wb)
            wb.save(output_file)
            logging.info(f"Unified report saved: {output_file} ({len(unified_summary)} devices across {len(family_files)} template families)")
            return processed_all
        finally:
            try:
                shutil.rmtree(tempdir, ignore_errors=True)
            except Exception:
                pass

    # ------------------------------------------------------------------
    # Packing slip workbook
    # ------------------------------------------------------------------

    def build_packing_slip_workbook(self, processed_data: Dict[str, Any], ip_list: List[str], customer: str, project: str, customer_po: str, sales_order: str, save_folder: str) -> str:
        """Generate a packing-slip Excel workbook from *processed_data*.

        Copies the packing slip template, populates a per-device sheet for
        each entry in *processed_data*, writes a summary sheet, then saves
        the finished file to *save_folder*.

        Args:
            processed_data: Dict keyed by IP/device ID with DataFrame values.
            ip_list: Ordered list of IPs for the summary table.
            customer: Customer name for the header.
            project: Project name for the header.
            customer_po: Purchase order number.
            sales_order: Sales order number.
            save_folder: Directory in which to write the output file.

        Returns:
            Absolute path to the saved packing slip file.

        Raises:
            FileNotFoundError: If the packing slip template does not exist.
        """
        packing_template_path = self.packing_slip_template
        logging.info(f"Loading packing slip template: {packing_template_path}")

        if not os.path.exists(packing_template_path):
            raise FileNotFoundError(f"Packing slip template not found at: {packing_template_path}")

        # Use secure temporary file instead of hardcoded path in cwd to prevent race conditions and symlink attacks
        temp_fd, temp_packing_slip = tempfile.mkstemp(prefix="PackingSlip_", suffix=".xlsx", dir=None)
        try:
            os.close(temp_fd)  # Close the file descriptor; we'll use the path
            shutil.copy(packing_template_path, temp_packing_slip)
        except Exception:
            try:
                os.unlink(temp_packing_slip)
            except Exception:
                pass
            raise

        try:
            wb_final = openpyxl.load_workbook(temp_packing_slip)

            # Find the summary sheet and the device template sheet dynamically
            # so the template file can use any sheet name.
            summary_sheet = None
            template_sheet = None
            for name in wb_final.sheetnames:
                if "summary" in name.lower():
                    summary_sheet = wb_final[name]
                else:
                    template_sheet = wb_final[name]

            if not summary_sheet:
                raise ValueError(f"No summary sheet found. Available: {wb_final.sheetnames}")
            if not template_sheet:
                raise ValueError(f"No device template sheet found. Available: {wb_final.sheetnames}")

            timestamp = datetime.now().strftime('%Y-%m-%d')
            safe_customer = re.sub(r'[^a-zA-Z0-9_]', '_', (customer or "Unknown").strip())
            safe_project = re.sub(r'[^a-zA-Z0-9_]', '_', (project or "Unknown").strip())

            filename = f"PackingSlip_{safe_customer}_{safe_project}_{timestamp}.xlsx"
            resolved_folder = os.path.realpath(save_folder)
            save_path = os.path.realpath(os.path.join(resolved_folder, filename))
            if not save_path.startswith(resolved_folder + os.sep) and save_path != resolved_folder:
                raise ValueError(f"Resolved save path escapes the target folder: {save_path}")

            logging.info(f"Packing slips will be saved as: {save_path}")

            # Timestamp in A2 — fill/title block applied later by _format_packing_slip_summary.
            capture_time = datetime.now().strftime('%Y-%m-%d @ %H:%M:%S')
            summary_sheet["A2"] = f"Capture Time = {capture_time}"
            summary_sheet["A2"].font = Font(color="00FF00", bold=True)

            ordered_items = sorted(processed_data.items(), key=lambda item: extract_ip_sort_key(item[0]))
            # Data rows start at row 7 (after the merged label rows 5-6).
            summary_start_row = 7
            summary_rows = []

            logging.info(f"Generating packing slips for {len(processed_data)} device(s)")

            for seq, (ip, device_data) in enumerate(ordered_items, start=1):
                if device_data.empty:
                    logging.warning(f"No data for {ip} - skipping")
                    continue

                logging.info(f"Columns for {ip}: {list(device_data.columns)}")

                try:
                    device_name = "Unknown_Device"
                    # Priority order: exact phrase match before broad "name" substring.
                    for pattern in ("system name", "device name", "hostname"):
                        for col in device_data.columns:
                            if pattern in col.lower():
                                device_name = str(device_data.iloc[0].get(col, "Unknown_Device"))
                                break
                        if device_name != "Unknown_Device":
                            break
                    if device_name == "Unknown_Device":
                        for col in device_data.columns:
                            if "name" in col.lower():
                                device_name = str(device_data.iloc[0].get(col, "Unknown_Device"))
                                break

                    device_name_clean = re.sub(r'[^a-zA-Z0-9_]', '_', device_name.strip())[:31]
                    if not device_name_clean:
                        device_name_clean = f"Device_{seq}"

                    new_sheet = wb_final.copy_worksheet(template_sheet)
                    new_sheet.title = device_name_clean

                    # Add quick navigation back to summary for easier review workflow.
                    new_sheet["A1"] = "Return"
                    new_sheet["A1"].hyperlink = f"#'{summary_sheet.title}'!A1"
                    new_sheet["A1"].style = "Hyperlink"

                    # NOTE: copy_worksheet already copies merged cell ranges;
                    # do NOT re-apply them or openpyxl raises ValueError.
                    # Labels (Customer:, Project:, Device ID:) and column headers
                    # at row 14 are already present in the template — only write values.
                    new_sheet["C5"] = customer or ""
                    new_sheet["C6"] = project or ""
                    new_sheet["C7"] = device_name

                    # SO and PO are written once at fixed cells B15 / C15
                    new_sheet["B15"] = sales_order or ""
                    new_sheet["C15"] = customer_po or ""

                    start_row = 15
                    for idx, row_dict in enumerate(device_data.to_dict('records')):
                        row_num = start_row + idx

                        # Case-insensitive column lookup so files with any casing work
                        row_lower = {k.lower(): v for k, v in row_dict.items()}
                        part_number = str(row_lower.get("part number", "")).strip()
                        serial_number = str(row_lower.get("serial number", "")).strip()
                        description = str(row_lower.get("description", "")).strip()

                        if not part_number:
                            part_number = str(row_lower.get("model number", "")).strip()

                        if part_number.lower() in ("nan", ""):
                            part_number = ""
                        if serial_number.lower() in ("nan", ""):
                            serial_number = ""
                        if description.lower() in ("nan", ""):
                            description = ""

                        if part_number or serial_number or description:
                            new_sheet[f"B{row_num}"] = sales_order or ""
                            new_sheet[f"C{row_num}"] = customer_po or ""
                            new_sheet[f"D{row_num}"] = self._sanitize_cell(part_number)
                            new_sheet[f"E{row_num}"] = self._sanitize_cell(serial_number)
                            new_sheet[f"F{row_num}"] = self._sanitize_cell(description)

                    self.autosize_sheet_columns(new_sheet)

                    summary_rows.append((seq, ip, device_name, new_sheet.title))

                    logging.debug(f"Written {len(device_data)} line items for device '{device_name}' with SO/PO")
                except Exception as exc:
                    logging.error(f"Error creating sheet for {ip}: {exc}")

            # New summary template columns: B=Customer, D=Project, F=IP Address, H=Device ID
            for row_offset, (seq, ip, device_name, sheet_title) in enumerate(summary_rows):
                row_num = summary_start_row + row_offset
                summary_sheet[f"B{row_num}"] = customer or ""
                summary_sheet[f"D{row_num}"] = project or ""
                summary_sheet[f"F{row_num}"] = str(ip)
                summary_sheet[f"H{row_num}"] = str(device_name)
                summary_sheet[f"H{row_num}"].hyperlink = f"#'{sheet_title}'!A1"
                summary_sheet[f"H{row_num}"].style = "Hyperlink"

            self._format_packing_slip_summary(summary_sheet, len(summary_rows))
            self.autosize_sheet_columns(summary_sheet)

            if summary_sheet.title in wb_final.sheetnames:
                idx = wb_final.sheetnames.index(summary_sheet.title)
                wb_final._sheets.insert(0, wb_final._sheets.pop(idx))

            if template_sheet.title in wb_final.sheetnames:
                del wb_final[template_sheet.title]

            self.autosize_workbook_columns(wb_final)
            wb_final.save(save_path)
            logging.info(f"Packing slips saved successfully: {save_path}")
            return save_path
        finally:
            if os.path.exists(temp_packing_slip):
                try:
                    os.remove(temp_packing_slip)
                except OSError:
                    pass

    # ------------------------------------------------------------------
    # Per-family packing slip support
    # ------------------------------------------------------------------

    @staticmethod
    def _detect_family_from_dataframe(df: Any) -> str:
        """Best-effort detect device family from an uploaded packing-slip dataframe.

        Looks at any column containing 'system type' or 'type' for known
        family signatures. Falls back to 'default'.
        """
        try:
            for col in df.columns:
                col_l = str(col).lower()
                if "type" not in col_l:
                    continue
                values = " ".join(str(v) for v in df[col].dropna().astype(str).head(10)).lower()
                if "rls" in values or "ciena" in values:
                    return "rls"
                if "psi" in values or "1830" in values or "nokia" in values:
                    return "psi"
        except Exception:
            pass
        return "default"

    def _populate_default_packing_slip_sheet(
        self,
        new_sheet: Any,
        device_data: Any,
        customer: str,
        project: str,
        customer_po: str,
        sales_order: str,
        device_name: str,
    ) -> None:
        """Default packing slip layout: B=SO, C=PO, D=PartNum, E=Serial, F=Description from row 15."""
        new_sheet["C5"] = customer or ""
        new_sheet["C6"] = project or ""
        new_sheet["C7"] = device_name
        new_sheet["B15"] = sales_order or ""
        new_sheet["C15"] = customer_po or ""

        start_row = 15
        for idx, row_dict in enumerate(device_data.to_dict("records")):
            row_num = start_row + idx
            row_lower = {str(k).lower(): v for k, v in row_dict.items()}
            part_number = str(row_lower.get("part number", "")).strip()
            serial_number = str(row_lower.get("serial number", "")).strip()
            description = str(row_lower.get("description", "")).strip()
            if not part_number:
                part_number = str(row_lower.get("model number", "")).strip()
            if part_number.lower() in ("nan", ""):
                part_number = ""
            if serial_number.lower() in ("nan", ""):
                serial_number = ""
            if description.lower() in ("nan", ""):
                description = ""
            if part_number or serial_number or description:
                new_sheet[f"B{row_num}"] = sales_order or ""
                new_sheet[f"C{row_num}"] = customer_po or ""
                new_sheet[f"D{row_num}"] = self._sanitize_cell(part_number)
                new_sheet[f"E{row_num}"] = self._sanitize_cell(serial_number)
                new_sheet[f"F{row_num}"] = self._sanitize_cell(description)

    def _populate_report_layout_packing_slip_sheet(
        self,
        new_sheet: Any,
        device_data: Any,
        customer: str,
        project: str,
        customer_po: str,
        sales_order: str,
        device_name: str,
        source_ip: str,
    ) -> None:
        """RLS/PSI report-template layout packing slip: SO/PO in C7 header, equipment in B-F from row 15.

        Columns: B=Slot/Port, C=Part Type, D=Part Number, E=Serial Number, F=Description.
        """
        new_sheet["C5"] = customer or ""
        new_sheet["C6"] = project or ""
        # Combined SO/PO in C7 (header label was set to "Sales Order / PO:" when template was created).
        so = (sales_order or "").strip()
        po = (customer_po or "").strip()
        if so and po:
            new_sheet["C7"] = f"SO {so} / PO {po}"
        elif so:
            new_sheet["C7"] = f"SO {so}"
        elif po:
            new_sheet["C7"] = f"PO {po}"
        # Right-side header block
        new_sheet["F5"] = source_ip or ""
        new_sheet["F6"] = device_name

        start_row = 15
        for idx, row_dict in enumerate(device_data.to_dict("records")):
            row_num = start_row + idx
            row_lower = {str(k).lower(): v for k, v in row_dict.items()}

            def pick(*keys: str) -> str:
                for k in keys:
                    v = row_lower.get(k)
                    if v is None:
                        continue
                    s = str(v).strip()
                    if s and s.lower() != "nan":
                        return s
                return ""

            slot = pick("slot/port", "slot", "slot/port number", "name")
            part_type = pick("part type", "type")
            part_number = pick("part number", "model number")
            serial_number = pick("serial number", "serial")
            description = pick("description")

            if slot or part_type or part_number or serial_number or description:
                new_sheet[f"B{row_num}"] = self._sanitize_cell(slot)
                new_sheet[f"C{row_num}"] = self._sanitize_cell(part_type)
                new_sheet[f"D{row_num}"] = self._sanitize_cell(part_number)
                new_sheet[f"E{row_num}"] = self._sanitize_cell(serial_number)
                new_sheet[f"F{row_num}"] = self._sanitize_cell(description)

    def build_unified_packing_slip_workbook(
        self,
        processed_data: Dict[str, Any],
        ip_list: List[str],
        customer: str,
        project: str,
        customer_po: str,
        sales_order: str,
        save_folder: str,
        family_for_ip: Optional[Dict[str, str]] = None,
        rls_packing_slip_template: Optional[str] = None,
        psi_packing_slip_template: Optional[str] = None,
    ) -> str:
        """Generate a packing slip workbook with per-device-family templates.

        For each device, picks the appropriate template (default / RLS / PSI)
        and merges all device sheets into one workbook with a unified Summary.
        Falls back to ``build_packing_slip_workbook`` when only the default
        family is needed (preserves byte-identical legacy output).
        """
        family_for_ip = family_for_ip or {}
        # Resolve family per IP, falling back to detection from the dataframe contents.
        resolved_family: Dict[str, str] = {}
        for ip, df in processed_data.items():
            fam = family_for_ip.get(ip)
            if not fam or fam == "default":
                fam = self._detect_family_from_dataframe(df)
            resolved_family[ip] = fam

        # If everything is default, use the legacy single-template path unchanged.
        if all(f == "default" for f in resolved_family.values()):
            return self.build_packing_slip_workbook(
                processed_data, ip_list, customer, project, customer_po, sales_order, save_folder,
            )

        templates = {
            "default": self.packing_slip_template,
            "rls": rls_packing_slip_template,
            "psi": psi_packing_slip_template,
        }
        for fam, path in templates.items():
            if not path or not os.path.exists(path):
                logging.warning(f"Packing slip template for family '{fam}' missing: {path} - using default")
                templates[fam] = self.packing_slip_template

        # Build base workbook from default template (Summary + Sheet1 device template).
        temp_fd, temp_path = tempfile.mkstemp(prefix="UnifiedPackingSlip_", suffix=".xlsx", dir=None)
        try:
            os.close(temp_fd)
            shutil.copy(self.packing_slip_template, temp_path)
        except Exception:
            try:
                os.unlink(temp_path)
            except Exception:
                pass
            raise

        try:
            wb_final = openpyxl.load_workbook(temp_path)

            summary_sheet = None
            default_template_sheet = None
            for name in wb_final.sheetnames:
                if "summary" in name.lower():
                    summary_sheet = wb_final[name]
                else:
                    default_template_sheet = wb_final[name]
            if not summary_sheet or not default_template_sheet:
                raise ValueError(f"Default packing slip template missing required sheets: {wb_final.sheetnames}")

            # Pre-load family-specific template sheets and stage as hidden sheets we copy from.
            family_template_sheets: Dict[str, Any] = {"default": default_template_sheet}
            staged_titles: List[str] = []
            for fam in ("rls", "psi"):
                tpath = templates[fam]
                if tpath == self.packing_slip_template:
                    family_template_sheets[fam] = default_template_sheet
                    continue
                src_wb = openpyxl.load_workbook(tpath)
                src_ws = None
                for n in src_wb.sheetnames:
                    if "summary" not in n.lower():
                        src_ws = src_wb[n]
                        break
                if not src_ws:
                    family_template_sheets[fam] = default_template_sheet
                    src_wb.close()
                    continue
                staged_title = self._unique_title(wb_final, f"_TPL_{fam}")
                self.copy_sheet(src_ws, wb_final, staged_title)
                family_template_sheets[fam] = wb_final[staged_title]
                staged_titles.append(staged_title)
                src_wb.close()

            timestamp = datetime.now().strftime("%Y-%m-%d")
            safe_customer = re.sub(r"[^a-zA-Z0-9_]", "_", (customer or "Unknown").strip())
            safe_project = re.sub(r"[^a-zA-Z0-9_]", "_", (project or "Unknown").strip())
            filename = f"PackingSlip_{safe_customer}_{safe_project}_{timestamp}.xlsx"
            resolved_folder = os.path.realpath(save_folder)
            save_path = os.path.realpath(os.path.join(resolved_folder, filename))
            if not save_path.startswith(resolved_folder + os.sep) and save_path != resolved_folder:
                raise ValueError(f"Resolved save path escapes the target folder: {save_path}")

            capture_time = datetime.now().strftime("%Y-%m-%d @ %H:%M:%S")
            summary_sheet["A2"] = f"Capture Time = {capture_time}"
            summary_sheet["A2"].font = Font(color="00FF00", bold=True)

            ordered_items = sorted(processed_data.items(), key=lambda item: extract_ip_sort_key(item[0]))
            summary_start_row = 7
            summary_rows: List[Tuple[int, str, str, str]] = []

            for seq, (ip, device_data) in enumerate(ordered_items, start=1):
                if device_data.empty:
                    continue
                fam = resolved_family.get(ip, "default")
                src_template = family_template_sheets.get(fam, default_template_sheet)

                device_name = "Unknown_Device"
                for pattern in ("system name", "device name", "hostname"):
                    for col in device_data.columns:
                        if pattern in str(col).lower():
                            device_name = str(device_data.iloc[0].get(col, "Unknown_Device"))
                            break
                    if device_name != "Unknown_Device":
                        break
                if device_name == "Unknown_Device":
                    for col in device_data.columns:
                        if "name" in str(col).lower():
                            device_name = str(device_data.iloc[0].get(col, "Unknown_Device"))
                            break

                device_name_clean = re.sub(r"[^a-zA-Z0-9_]", "_", device_name.strip())[:31]
                if not device_name_clean:
                    device_name_clean = f"Device_{seq}"
                device_name_clean = self._unique_title(wb_final, device_name_clean)

                # Copy template sheet (handles same-workbook copy via copy_worksheet
                # for default; cross-style merges already staged for rls/psi).
                if src_template.parent is wb_final:
                    new_sheet = wb_final.copy_worksheet(src_template)
                    new_sheet.title = device_name_clean
                else:
                    new_sheet = self.copy_sheet(src_template, wb_final, device_name_clean)

                new_sheet["A1"] = "Return"
                new_sheet["A1"].hyperlink = f"#'{summary_sheet.title}'!A1"
                new_sheet["A1"].style = "Hyperlink"

                if fam in ("rls", "psi"):
                    self._populate_report_layout_packing_slip_sheet(
                        new_sheet, device_data, customer, project, customer_po, sales_order, device_name, str(ip),
                    )
                else:
                    self._populate_default_packing_slip_sheet(
                        new_sheet, device_data, customer, project, customer_po, sales_order, device_name,
                    )

                self.autosize_sheet_columns(new_sheet)
                summary_rows.append((seq, str(ip), device_name, new_sheet.title))

            # Populate unified Summary table.
            for row_offset, (seq, ip, device_name, sheet_title) in enumerate(summary_rows):
                row_num = summary_start_row + row_offset
                summary_sheet[f"B{row_num}"] = customer or ""
                summary_sheet[f"D{row_num}"] = project or ""
                summary_sheet[f"F{row_num}"] = ip
                summary_sheet[f"H{row_num}"] = device_name
                summary_sheet[f"H{row_num}"].hyperlink = f"#'{sheet_title}'!A1"
                summary_sheet[f"H{row_num}"].style = "Hyperlink"

            self._format_packing_slip_summary(summary_sheet, len(summary_rows))
            self.autosize_sheet_columns(summary_sheet)

            # Move Summary to the front and remove staged template sheets.
            if summary_sheet.title in wb_final.sheetnames:
                idx = wb_final.sheetnames.index(summary_sheet.title)
                wb_final._sheets.insert(0, wb_final._sheets.pop(idx))
            for staged in staged_titles + [default_template_sheet.title]:
                if staged in wb_final.sheetnames:
                    del wb_final[staged]

            self.autosize_workbook_columns(wb_final)
            wb_final.save(save_path)
            logging.info(f"Unified per-family packing slip saved: {save_path}")
            return save_path
        finally:
            if os.path.exists(temp_path):
                try:
                    os.remove(temp_path)
                except OSError:
                    pass

    @staticmethod
    def _unique_title(wb: Any, base: str) -> str:
        """Return a unique sheet title in *wb* derived from *base* (Excel 31-char limit)."""
        title = base[:31]
        if title not in wb.sheetnames:
            return title
        i = 2
        while True:
            suffix = f"_{i}"
            candidate = (base[: 31 - len(suffix)] + suffix)
            if candidate not in wb.sheetnames:
                return candidate
            i += 1
