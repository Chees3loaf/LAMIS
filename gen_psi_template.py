# -*- coding: utf-8 -*-
"""Generate Nokia_PSI_Report_Template.xlsx — run once to create/refresh the file."""

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BLUE  = "FF156082"
WHITE = "FFFFFFFF"
THIN  = Side(border_style="thin")


def blue_fill():
    return PatternFill(fill_type="solid", fgColor=BLUE)


def white_fill():
    return PatternFill(fill_type="solid", fgColor=WHITE)


def blue_font(size=10):
    return Font(name="Calibri", bold=True, size=size, color=WHITE)


def plain_font(size=10):
    return Font(name="Calibri", bold=False, size=size)


def thin_border():
    return Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def apply_section_header(ws, coord, value, merge_range=None, size=10, row_height=16):
    if merge_range:
        ws.merge_cells(merge_range)
    cell = ws[coord]
    cell.value = value
    cell.fill = blue_fill()
    cell.font = blue_font(size)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border()
    row_num = int(''.join(filter(str.isdigit, coord)))
    ws.row_dimensions[row_num].height = row_height


def apply_col_header(ws, coord, value, size=8, wrap=False):
    cell = ws[coord]
    cell.value = value
    cell.fill = blue_fill()
    cell.font = blue_font(size)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=wrap)
    cell.border = thin_border()


def apply_data_row(ws, row):
    for col in "BCDEF":
        c = ws[f"{col}{row}"]
        c.fill = white_fill()
        c.font = plain_font(11)
        c.border = thin_border()
        c.alignment = Alignment(horizontal="center", vertical="center")


# ── Build from existing template to preserve theme/style ─────────────────────
wb = load_workbook("data/Device_Report_Template.xlsx")
ws = wb.active
ws.title = "Customer-project"

# Extend equipment data rows to 40 (rows 15–54)
for r in range(15, 55):
    apply_data_row(ws, r)

# Clear old end sentinel
ws["E93"] = None

# Gap rows 55–57 — blank, no border
for r in range(55, 58):
    for col in "BCDEF":
        c = ws[f"{col}{r}"]
        c.fill = white_fill()
        c.border = Border()
        c.value = None

# ── BLUE DIVIDER rows 58–59 ───────────────────────────────────────────────────
ws.merge_cells("B58:F59")
cell = ws["B58"]
cell.value = "ADDITIONAL NODE INFORMATION"
cell.fill = blue_fill()
cell.font = blue_font(14)
cell.alignment = Alignment(horizontal="center", vertical="center")
cell.border = thin_border()
ws.row_dimensions[58].height = 20
ws.row_dimensions[59].height = 20

# ── SOFTWARE INFORMATION (rows 60–71) ────────────────────────────────────────
apply_section_header(ws, "B60", "SOFTWARE INFORMATION \u2014 show software dynamic",
                     merge_range="B60:F60")
for col, label in zip("BCDEF", ["NAME", "RELEASE VERSION", "RPMS LOADED", "RPMS TOTAL", "NOTES"]):
    apply_col_header(ws, f"{col}61", label)
for r in range(62, 72):
    apply_data_row(ws, r)

# ── SLOT STATUS (rows 72–103) ─────────────────────────────────────────────────
apply_section_header(ws, "B72", "SLOT STATUS \u2014 show slot *",
                     merge_range="B72:F72")
for col, label in zip("BCDEF", ["SLOT", "PROGRAMMED TYPE", "PRESENT TYPE", "ADMIN STATE", "OPER STATE / QUALIFIER"]):
    apply_col_header(ws, f"{col}73", label, wrap=True)
for r in range(74, 104):
    apply_data_row(ws, r)

# ── REDUNDANCY INFORMATION (rows 104–115) ────────────────────────────────────
apply_section_header(ws, "B104", "REDUNDANCY INFORMATION \u2014 show redundancy 1 detail",
                     merge_range="B104:F104")
for col, label in zip("BCDEF", ["NAME", "CLOCK SWITCH", "EC SELECTION", "STATE", "NOTES"]):
    apply_col_header(ws, f"{col}105", label)
for r in range(106, 116):
    apply_data_row(ws, r)

# ── POWER FEED STATUS (rows 116–127) ─────────────────────────────────────────
apply_section_header(ws, "B116", "POWER FEED STATUS \u2014 show pf *",
                     merge_range="B116:F116")
for col, label in zip("BCDEF", ["SLOT", "TYPE", "ADMIN STATE", "OPER STATE", "NOTES"]):
    apply_col_header(ws, f"{col}117", label)
for r in range(118, 128):
    apply_data_row(ws, r)

# ── INTERFACE TOPOLOGY (rows 128–169) ────────────────────────────────────────
apply_section_header(ws, "B128", "INTERFACE TOPOLOGY \u2014 show interface topology *",
                     merge_range="B128:F128")
for col, label in zip("BCDEF", ["INTERFACE", "INTERFACE TYPE", "CONNECTED TO", "TYPE FROM", "NOTES"]):
    apply_col_header(ws, f"{col}129", label)
for r in range(130, 170):
    apply_data_row(ws, r)

# End sentinel keeps file height consistent
ws["E170"] = " "

wb.save("data/Nokia_PSI_Report_Template_new.xlsx")
print("Saved: data/Nokia_PSI_Report_Template_new.xlsx  (rows:", ws.max_row, ")")
