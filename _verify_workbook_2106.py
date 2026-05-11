from pathlib import Path
from openpyxl import load_workbook

wb_path = Path(r"C:\Users\ZackerySimino\OneDrive - LightRiver Technologies Inc\Desktop\Inventory and Packing slips\BPA\TO3\Master_BPA_Task_Order_3_Raw_Report_2026-05-10_21-06.xlsx")

if not wb_path.exists():
    raise SystemExit(f"Workbook not found: {wb_path}")

wb = load_workbook(wb_path, data_only=True)

print(f"Workbook: {wb_path.name}")
print(f"Total sheets: {len(wb.sheetnames)}")
print(f"Device sheets (excluding Summary): {len([s for s in wb.sheetnames if s != 'Summary'])}")

sheet_name = "CUSI002_7705"
ws = wb[sheet_name]

print("\nCUSI002_7705 rows containing 3HE06792AA:")
found = False
for r_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
    if any((v is not None and "3HE06792AA" in str(v)) for v in row):
        found = True
        print(f"Row {r_idx}: {row}")
if not found:
    print("None")

expected_desc = "Fan Module (SAR-8 Shelf V2) Ext Temp"
expected_found = False
for sname in wb.sheetnames:
    if sname == "Summary":
        continue
    ws2 = wb[sname]
    for row in ws2.iter_rows(values_only=True):
        if any(v == expected_desc for v in row if v is not None):
            expected_found = True
            break
    if expected_found:
        break
print(f"\nExpected fan description present anywhere: {expected_found}")

not_found_cells = 0
for sname in wb.sheetnames:
    if sname == "Summary":
        continue
    ws2 = wb[sname]
    for row in ws2.iter_rows(values_only=True):
        for v in row:
            if v is not None and str(v).strip() == "Not Found":
                not_found_cells += 1

print(f"Literal 'Not Found' cell count (all columns): {not_found_cells}")
