"""Unit tests for gui/workbook_builder.py — sheet utilities and data combination."""

import unittest
from unittest.mock import MagicMock, patch
import pandas as pd
import openpyxl

from gui.workbook_builder import WorkbookBuilder


def _make_builder() -> WorkbookBuilder:
    """Return a WorkbookBuilder with mocked dependencies."""
    db_cache = MagicMock()
    db_cache.db_path = ":memory:"
    return WorkbookBuilder(db_cache=db_cache, template_path="", packing_slip_template="")


class TestAutosizeSheetColumns(unittest.TestCase):

    def test_empty_sheet_does_not_raise(self):
        builder = _make_builder()
        wb = openpyxl.Workbook()
        ws = wb.active
        builder.autosize_sheet_columns(ws)  # should not raise

    def test_column_width_set_for_populated_cell(self):
        builder = _make_builder()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "Hello World"  # 11 chars
        builder.autosize_sheet_columns(ws, min_width=5, max_width=50)
        # width = max(11 + 2, 5) = 13
        self.assertEqual(ws.column_dimensions["A"].width, 13)

    def test_min_width_floor_respected(self):
        builder = _make_builder()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "Hi"  # 2 chars → 2+2=4, below min_width=10
        builder.autosize_sheet_columns(ws, min_width=10, max_width=50)
        self.assertEqual(ws.column_dimensions["A"].width, 10)

    def test_max_width_ceiling_respected(self):
        builder = _make_builder()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "x" * 200  # way over any max
        builder.autosize_sheet_columns(ws, min_width=5, max_width=30)
        self.assertEqual(ws.column_dimensions["A"].width, 30)

    def test_none_cell_values_skipped(self):
        builder = _make_builder()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = None
        ws["A2"] = "Test"  # 4 chars → width = max(6, 10) = 10
        builder.autosize_sheet_columns(ws, min_width=10, max_width=50)
        self.assertEqual(ws.column_dimensions["A"].width, 10)


class TestCopySheet(unittest.TestCase):

    def test_copies_cell_values(self):
        builder = _make_builder()
        src_wb = openpyxl.Workbook()
        src_ws = src_wb.active
        src_ws["A1"] = "alpha"
        src_ws["B2"] = 42

        dst_wb = openpyxl.Workbook()
        result_ws = builder.copy_sheet(src_ws, dst_wb, "Copied")
        self.assertEqual(result_ws["A1"].value, "alpha")
        self.assertEqual(result_ws["B2"].value, 42)

    def test_new_sheet_name_used(self):
        builder = _make_builder()
        src_wb = openpyxl.Workbook()
        src_ws = src_wb.active
        dst_wb = openpyxl.Workbook()
        result_ws = builder.copy_sheet(src_ws, dst_wb, "MySheet")
        self.assertEqual(result_ws.title, "MySheet")

    def test_sheet_added_to_target_workbook(self):
        builder = _make_builder()
        src_wb = openpyxl.Workbook()
        dst_wb = openpyxl.Workbook()
        builder.copy_sheet(src_wb.active, dst_wb, "Extra")
        self.assertIn("Extra", dst_wb.sheetnames)


class TestCombineAndFormatData(unittest.TestCase):

    def _make_ip_df(self, ip: str, rows: int = 2) -> pd.DataFrame:
        return pd.DataFrame({
            "IP Address": [ip] * rows,
            "Name": [f"port-{i}" for i in range(rows)],
            "Part Number": [f"PN-{i}" for i in range(rows)],
            "Serial Number": [f"SN-{i}" for i in range(rows)],
            "Description": [f"Desc-{i}" for i in range(rows)],
        })

    def test_returns_dataframe(self):
        builder = _make_builder()
        data = {"10.0.0.1": self._make_ip_df("10.0.0.1")}
        result = builder.combine_and_format_data(data)
        self.assertIsInstance(result, pd.DataFrame)

    def test_empty_input_returns_empty_dataframe(self):
        builder = _make_builder()
        result = builder.combine_and_format_data({})
        self.assertIsInstance(result, pd.DataFrame)
        self.assertEqual(len(result), 0)

    def test_multiple_ips_all_rows_present(self):
        builder = _make_builder()
        data = {
            "10.0.0.1": self._make_ip_df("10.0.0.1", rows=3),
            "10.0.0.2": self._make_ip_df("10.0.0.2", rows=2),
        }
        result = builder.combine_and_format_data(data)
        self.assertEqual(len(result), 5)

    def test_ip_address_column_present(self):
        builder = _make_builder()
        data = {"10.0.0.1": self._make_ip_df("10.0.0.1")}
        result = builder.combine_and_format_data(data)
        self.assertIn("IP Address", result.columns)


class TestSetupSummarySheetHeader(unittest.TestCase):

    def _blank_ws(self) -> openpyxl.worksheet.worksheet.Worksheet:
        wb = openpyxl.Workbook()
        return wb.active

    def test_customer_written_to_b7(self):
        builder = _make_builder()
        ws = self._blank_ws()
        builder._setup_summary_sheet_header(ws, "Acme Corp", "Proj-X", ["10.0.0.1"])
        self.assertEqual(ws["B7"].value, "Acme Corp")

    def test_project_written_to_d7(self):
        builder = _make_builder()
        ws = self._blank_ws()
        builder._setup_summary_sheet_header(ws, "Acme", "Proj-X", [])
        self.assertEqual(ws["D7"].value, "Proj-X")

    def test_ip_list_joined_in_f7(self):
        builder = _make_builder()
        ws = self._blank_ws()
        builder._setup_summary_sheet_header(ws, "", "", ["10.0.0.1", "10.0.0.2"])
        self.assertEqual(ws["F7"].value, "10.0.0.1, 10.0.0.2")

    def test_empty_ip_list_f7_empty_string(self):
        builder = _make_builder()
        ws = self._blank_ws()
        builder._setup_summary_sheet_header(ws, "C", "P", [])
        self.assertEqual(ws["F7"].value, "")

    def test_capture_time_written_to_a2(self):
        builder = _make_builder()
        ws = self._blank_ws()
        builder._setup_summary_sheet_header(ws, "C", "P", [])
        self.assertIn("Capture Time", str(ws["A2"].value))


class TestPopulateSummaryTable(unittest.TestCase):

    def test_rows_written_at_correct_offset(self):
        builder = _make_builder()
        wb = openpyxl.Workbook()
        ws = wb.active
        items = [("10.0.0.1", "Router-A", "Sheet1")]
        builder._populate_summary_table(ws, items, start_row=10)
        self.assertEqual(ws["C10"].value, "10.0.0.1")
        self.assertEqual(ws["D10"].value, "Router-A")

    def test_multiple_items_sorted_by_ip(self):
        builder = _make_builder()
        wb = openpyxl.Workbook()
        ws = wb.active
        items = [
            ("10.0.0.20", "B", "SheetB"),
            ("10.0.0.3", "A", "SheetA"),
        ]
        builder._populate_summary_table(ws, items, start_row=10)
        # First row (row 10) should be the lower IP
        self.assertEqual(ws["C10"].value, "10.0.0.3")
        self.assertEqual(ws["C11"].value, "10.0.0.20")

    def test_row_number_increments(self):
        builder = _make_builder()
        wb = openpyxl.Workbook()
        ws = wb.active
        items = [("10.0.0.1", "A", "S1"), ("10.0.0.2", "B", "S2")]
        builder._populate_summary_table(ws, items, start_row=5)
        self.assertEqual(ws["B5"].value, 1)
        self.assertEqual(ws["B6"].value, 2)


if __name__ == "__main__":
    unittest.main()
